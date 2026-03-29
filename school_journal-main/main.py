import os, re
import datetime

from google_oauth_downloader import download_class_journals_oauth
from gsheet_reader import read_class_links_from_gsheet
import marks_dict as m
import generate_page as g
import helpers as h
from helpers import get_assistant_principal_for_class
from helpers import get_tutor_info_from_xlsx
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from helpers import get_tutor_info_from_xlsx
from generate_page import make_mark_value
# Глобальные переменные для тьютора
TutorName = ""
TutorText = ""

def _is_blank(x):
    return x is None or str(x).strip() == ""

def _norm(x):
    # нормализация пробелов/переносов и регистра
    s = str(x).replace("\r\n", "\n").replace("\xa0", " ")
    return " ".join(s.split())

class Student:
    def __init__(self, first_name="N/A", last_name="N/A", level=-1, subjects=None):
        self.first_name = first_name
        self.last_name = last_name
        self.level = level
        if subjects is None:
            self.subjects = []
        else:
            self.subjects = subjects


class Subject:
    def __init__(self, name="N/A", teacher="N/A", module=-1,
                 descriptor="N/A", marks=None,
                 comment=None, retake=None):
        self.name = name
        self.teacher = teacher
        self.module = module
        self.descriptor = descriptor
        self.marks = {} if marks is None else marks
        # новое
        self.comment = comment   # «Комментарий» по предмету
        self.retake = retake     # «Пересдача» по предмету (заполним позже)



def enrich_mark(mark_value):
    if isinstance(mark_value, (int, float)):
        for key, value in m.marks_dict.items():
            if value['bounds'][0] < mark_value <= value['bounds'][1]:
                return [value['order'], key, mark_value]
        return [-1, 'NOT_FOUND']
    else:
        try:
            result = [m.marks_dict[mark_value]['order'], mark_value]
            return result
        except KeyError:
            print(f'No key in the mark dictionary for {mark_value}')
            if mark_value == '-':
                return [-1, 'NOT_FOUND']
            else:
                return [-2, mark_value]



def write_error_log(errors, file_path):
    if errors:
        with open(file_path, 'w', encoding='utf-8') as file:
            for item in errors:
                current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                file.write(f"{current_time} - {item}\n")


def create_students_from_xlsx(file_path):
    global TutorName, TutorText
    wb = load_workbook(file_path, data_only=True)
    students = []

    for sheet in wb.sheetnames:
        # print(sheet)
        title_norm = sheet.lower()
        # === Особая обработка листа "Тьютор | Tutor"
        if title_norm == "тьютор | tutor" or "тьютор" in title_norm or "tutor" in title_norm:
            # Берём учителя из C2, «Дескриптор | Descriptor» — из C4 (как на обычных предметах)
            TutorName = wb[sheet]["C2"].value
            TutorText = wb[sheet]["C4"].value
            # Этот лист дальше не обрабатываем как предмет
            continue
        # === конец обработки листа тьютора

        current_sheet = wb[sheet]

        name_row_found = False
        criteria_row_found = False
        criteria_row_cell = None

        # индексы «служебных» колонок (определим, когда найдём шапку критериев)
        comment_col = None
        retake_col = None

        for row in current_sheet.iter_rows():
            # ожидаем: имя в колонке A, заголовок критериев — в колонке B
            name_cell = row[0]
            criteria_header_cell = row[1]

            # нормализация заголовка
            def _norm_local(s: str) -> str:
                s = s.replace("\r\n", "\n").replace("\xa0", " ")
                s = re.sub(r"[ \t]*\|\s*", " | ", s)   # пробелы вокруг «|»
                s = re.sub(r"[ \t]+", " ", s)          # схлоп пробелов
                return s.strip().lower()

            # нашли строку заголовка критериев
            if _norm_local(criteria_header_cell.value or "") == _norm_local("Критерии оценивания | \nAssessment criteria"):
                criteria_row_found = True
                criteria_row_cell = criteria_header_cell

                # Определим номера колонок «Комментарий» / «Пересдача» в ЭТОЙ строке
                hdr_row = criteria_row_cell.row
                comment_col = None
                retake_col = None
                for col in range(1, current_sheet.max_column + 1):
                    hdr_val = current_sheet.cell(row=hdr_row, column=col).value
                    hdr_str = (str(hdr_val) if hdr_val is not None else "").strip().lower()
                    if ("коммент" in hdr_str) or ("примеч" in hdr_str) or ("замеч" in hdr_str) or ("comment" in hdr_str) or ("note" in hdr_str):
                        comment_col = col
                    if ("пересда" in hdr_str) or ("retake" in hdr_str) or ("resit" in hdr_str) or ("make-up" in hdr_str):
                        retake_col = col

            # нашли строку заголовков имён
            if str(name_cell.value).strip().lower() == "имя":
                name_row_found = True

            # пошли строки данных учеников
            elif name_row_found and name_cell.value is not None:
                first_name = row[name_cell.column - 1].value  # колонка A
                last_name = row[name_cell.column].value       # колонка B
                level = str(current_sheet['C1'].value)

                # ищем уже созданного ученика (ФИ + класс)
                existing_student = None
                for st in students:
                    if st.first_name and st.last_name:
                        if st.first_name.lower() == str(first_name).lower() and \
                           st.last_name.lower() == str(last_name).lower() and \
                           st.level == level:
                            existing_student = st
                            break

                # создаём предмет для этого листа
                subject = Subject(
                    name=sheet,
                    teacher=current_sheet['C2'].value,
                    module=int(current_sheet['C3'].value),
                    descriptor=current_sheet['C4'].value
                )

                # собираем оценки
                marks = {}
                if criteria_row_found and criteria_row_cell is not None:
                    hdr_row = criteria_row_cell.row
                    first_col = criteria_row_cell.column + 1

                    for col in range(first_col, current_sheet.max_column + 1):
                        header = current_sheet.cell(row=hdr_row, column=col).value
                        if _is_blank(header):
                            continue

                        header_norm = _norm(header)
                        low = header_norm.lower()

                        # служебные колонки (не критерии)
                        if col == comment_col:
                            continue
                        if col == retake_col:
                            continue

                        # значение для текущего ученика
                        cell = current_sheet.cell(row=name_cell.row, column=col)
                        if _is_blank(cell.value):
                            # критерий существует, оценка отсутствует
                            marks[header_norm] = None
                        else:
                            marks[header_norm] = enrich_mark(cell.value)

                subject.marks = marks

                # Прочитаем комментарий / пересдачу из служебных колонок (если они есть)
                row_idx = name_cell.row
                if comment_col:
                    val = current_sheet.cell(row=row_idx, column=comment_col).value
                    if val:
                        subject.comment = str(val).strip()
                if retake_col:
                    val = current_sheet.cell(row=row_idx, column=retake_col).value
                    if val:
                        subject.retake = str(val).strip()

                # добавляем предмет ученику
                if existing_student is not None:
                    existing_student.subjects.append(subject)
                else:
                    student = Student(first_name, last_name, level)
                    student.subjects.append(subject)
                    students.append(student)

            # дошли до пустой строки после списка учеников — выходим с листа
            elif name_row_found and name_cell.value is None:
                break

    return students

from openpyxl import load_workbook



def fill_header(report_data, template_path, output_path, workbook_path) -> list:
    """
    Генерирует «первые страницы» по шаблону.
    Берёт общий текст и имя тьютора с листа «Тьютор | Tutor»,
    а также персональный комментарий ученика (если заполнен в колонке "Комментарий").
    """
    errors, paths = [], []
    tutor_cache: dict[str, tuple[str, str]] = {}  # class_code -> (tutor_text, tutor_name)


    for student in report_data:
        class_code = student.level or ""
        ap = get_assistant_principal_for_class(class_code) or ""

        # 1) Берём общий текст/имя тьютора (из глобалей или helpers)
        global TutorName, TutorText
        if TutorName or TutorText:
            tutor_name = TutorName or ""
            tutor_text = TutorText or ""
        else:
            if class_code not in tutor_cache:
                tutor_cache[class_code] = get_tutor_info_from_xlsx(
                    workbook_path, class_code, debug=False
                )
            tutor_text, tutor_name = tutor_cache[class_code]
        tutor_text_final = (tutor_text or "").strip()
        personal_comment = ""
        wb = load_workbook(workbook_path, data_only=True)
        for sheet in wb.sheetnames:
            title_norm = (sheet or "").strip().lower()
            if title_norm == "тьютор | tutor" or "тьютор" in title_norm or "tutor" in title_norm:
                ws = wb[sheet]
                target = student.last_name
                addr = ""
                # --- найти строку заголовков и номера нужных колонок ---
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                        min_col=1, max_col=ws.max_column):
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.strip().lower() == target.lower():
                            addr = cell.coordinate  # например: 'B9'
                            break
                    if addr:
                        break
                #target2 = student.first_name
                #addr2 = "A"+addr[1:]
                #if target2.lower() == wb[sheet][addr2].value.lower():
                    #personal_comment = wb[sheet]["C"+addr2[1:]].value
                ws = wb[sheet]

                def norm(s):
                    return " ".join(str(s or "").split()).strip().lower()

                # 1) ищем строку ученика в колонке B (Фамилия)
                row_idx = None
                for cell in ws['B']:  # ws['B'] -> кортеж всех ячеек колонки B
                    if norm(cell.value) == norm(student.last_name):
                        row_idx = cell.row
                        break

                # 2) берём комментарий из колонки C этой же строки
                personal_comment = ""
                if row_idx:
                    # если уверен, что комментарий в 'C'
                    personal_comment = ws[f"C{row_idx}"].value or ""
                    # или надёжнее по номеру колонки, если ты его вычислял как col_comment:
                    # personal_comment = ws.cell(row=row_idx, column=col_comment).value or ""

        tutor_text_final = (TutorText or "").strip()
        if personal_comment:
            personal_comment = personal_comment.strip()
            if personal_comment:
                tutor_text_final = tutor_text_final + ("\n\n" if tutor_text_final else "") + personal_comment

        # 3) Общий контекст для шаблона (всегда создаём ДО try)
        ctx = {
            "StudentName": f"{student.first_name} {student.last_name}",
            "Class": class_code,
            "AssistantPrincipal": ap,
            "TutorName": tutor_name or "",
            "TutorText": tutor_text_final or "",
            "HasTutor": bool(tutor_text_final),
        }

        # 4) Рендер/сохранение
        try:
            tpl = DocxTemplate(template_path)  # новый экземпляр на каждого ученика
            tpl.render(ctx)
            os.makedirs(output_path, exist_ok=True)
            filepath = os.path.join(output_path, f"Header {ctx['StudentName']}.docx")
            tpl.save(filepath)

            paths.append({
                "student": f"{ctx['StudentName']}{ctx['Class']}",
                "filepath": filepath,
                "filename": f"{ctx['StudentName']}.docx",
            })
        except Exception as e:
            errors.append(f"Some error occurred with {ctx['StudentName']}: {e}")
            break

    write_error_log(errors, "errors.txt")
    return paths



if __name__ == '__main__':
    
    # 0) пути
    script_dir = os.path.dirname(os.path.abspath(__file__))
    header_template = os.path.join(script_dir, 'input', 'first page template.docx')

    # 1) читаем словарь {класс: ссылка}
    GOOGLE_CLASSES_URL = os.getenv("GOOGLE_CLASSES_URL", "")
    if not GOOGLE_CLASSES_URL:
        raise RuntimeError("Set GOOGLE_CLASSES_URL in environment")
    class_links = read_class_links_from_gsheet(GOOGLE_CLASSES_URL)

    # 2) качаем все журналы по OAuth → input/downloads/journal_<класс>.xlsx
    downloads_dir = os.path.join(script_dir, 'input', 'downloads')
    downloaded_workbooks = download_class_journals_oauth(class_links, out_dir=downloads_dir)

    # 3) обрабатываем КАЖДЫЙ файл отдельно и складываем вывод в output/<Класс>/
    for workbook_path in downloaded_workbooks:
        # имя класса берём из имени файла: journal_XXXX.xlsx → XXXX
        base = os.path.splitext(os.path.basename(workbook_path))[0]
        class_name = base.replace('journal_', '', 1)

        # своя temp-папка и своя итоговая папка для класса
        output_temp_path = os.path.join(script_dir, 'output', 'temp', class_name)
        output_path = os.path.join(script_dir, 'output', class_name)
        os.makedirs(output_temp_path, exist_ok=True)
        os.makedirs(output_path, exist_ok=True)

        print(
            f"\n==================\nОбрабатываю класс: {class_name}\nФайл: {workbook_path}\nВывод: {output_path}\n==================")

        # обычный пайплайн на конкретный workbook
        students = create_students_from_xlsx(workbook_path)
        headers = fill_header(students, header_template, output_temp_path, workbook_path)
        class_tables = g.generate_subject(students, output_temp_path)

        # склейка: к титулу каждого ученика добавляем все его предметы; итог — в output/<Класс>/
        for header in headers:
            for t in class_tables:
                if header['student'] == t['student']:
                    h.merge_documents(header['filepath'], t['filepath'], header['filepath'])
            h.copy_and_rename_file(header['filepath'], output_path, header['filename'])

        # чистим временные файлы именно этого класса
        h.cleanup_folder(output_temp_path)

    print("\nГотово. Все журналы обработаны и разложены по папкам output/<Класс>/")

    clean_up_temp = True
    script_dir = os.path.dirname(os.path.abspath(__file__))
    header_template = os.path.join(script_dir, 'input/first page template.docx')
    workbook_path = os.path.join(script_dir, 'input/Journal.xlsx')
    output_temp_path = os.path.join(script_dir, 'output/temp')
    output_path = os.path.join(script_dir, 'output')

    students = create_students_from_xlsx(workbook_path)
    headers = fill_header(students, header_template, output_temp_path)
    class_tables = g.generate_subject(students, output_temp_path)

    # merging documents together
    for header in headers:
        for t in class_tables:
            if header['student'] == t['student']:
                h.merge_documents(header['filepath'], t['filepath'], header['filepath'])
        # copy the result into output table
        h.copy_and_rename_file(header['filepath'], output_path, header['filename'])

    # cleaning temp
    if clean_up_temp:
        h.cleanup_folder(output_temp_path)
    '''    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    header_template = os.path.join(script_dir, 'input/first page template.docx')
    workbook_path = os.path.join(script_dir, 'input/Journal4a.xlsx')
    output_temp_path = os.path.join(script_dir, 'output/temp')
    output_path = os.path.join(script_dir, 'output')

    students = create_students_from_xlsx(workbook_path)

    # headers = fill_header(students, header_template, output_temp_path, workbook_path)
    headers = fill_header(students, header_template, output_temp_path, workbook_path)


    class_tables = g.generate_subject(students, output_temp_path)

    # merging documents together
    for header in headers:
        for t in class_tables:
            if header['student'] == t['student']:
                h.merge_documents(header['filepath'], t['filepath'], header['filepath'])
        # copy the result into output table
        h.copy_and_rename_file(header['filepath'], output_path, header['filename'])

    # cleaning temp
    if clean_up_temp:
        h.cleanup_folder(output_temp_path)
    '''
    # DEBUG - display student data as a tree
    # for s in students:
    # 	print(f'{s.first_name} {s.last_name} {s.level}')
    # 	for sub in s.subjects:
    # 		print(f'   {sub.name} {sub.teacher} {sub.descriptor}')
    # 		for key, value in sub.marks.items():
    # 			print(f'	  {key} {value}')

    # DEBUG - display created files
    # for h in headers:
    #     print(h)
    # print('')
    # for t in class_tables:
    #     print(t)
