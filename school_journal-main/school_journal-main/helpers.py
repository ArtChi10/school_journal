import shutil
import os
from docx import Document
from docxcompose.composer import Composer
import re

def merge_documents(template1_path, template2_path, output_path):
    doc1 = Document(template1_path)
    doc2 = Document(template2_path)
    doc1.add_page_break()
    composer = Composer(doc1)
    composer.append(doc2)
    composer.save(output_path)


def copy_and_rename_file(src_path, dest_dir, new_name):
    """
    Copy a file from src_path to dest_dir and rename it to new_name.

    :param src_path: Path to the source file
    :param dest_dir: Directory where the file should be copied to
    :param new_name: New name for the copied file
    """
    # Ensure the destination directory exists
    dest_dir = dest_dir.rstrip('/')  # Remove any trailing slash for consistency
    
    # Construct the destination path
    dest_path = f"{dest_dir}/{new_name}"
    
    # Copy the file to the new location with the new name
    shutil.copy(src_path, dest_path)


def cleanup_folder(folder_path):
    """
    Remove all files and subdirectories in the specified folder.

    :param folder_path: Path to the folder to clean up
    """
    # Ensure the folder exists
    if not os.path.exists(folder_path):
        print(f"The folder {folder_path} does not exist.")
        return
    
    # Iterate over the contents of the folder and remove each item
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)  # Remove file or link
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)  # Remove directory and its contents
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")

# Запрещённые для Windows символы в именах файлов: <>:"/\|?*
_INVALID_CHARS_RE = re.compile(r'[<>:"/\\|?*]+')

def sanitize_filename(name: str, repl: str = "—") -> str:
    """
    Делает имя файла безопасным для Windows/macOS/Linux:
    - заменяет запрещённые символы на repl (по умолчанию — длинное тире),
    - обрезает пробелы/точки по краям,
    - предотвращает пустые имена.
    """
    s = str(name) if name is not None else ""
    s = _INVALID_CHARS_RE.sub(repl, s)
    s = s.strip().strip(".")
    return s or "untitled"


def ensure_dir(path: str) -> None:
    """Гарантирует существование директории."""
    os.makedirs(path, exist_ok=True)

# --- Assistant Principals (Завучи) -----------------------------------------

# Нормализация записи класса (снимаем пробелы/регистр, приводим K/К и C/С к единому виду)
def _normalize_class(code: str) -> str:
    if code is None:
        return ""
    s = str(code).strip().upper().replace(" ", "")
    # приводим кириллицу к латинице для двусмысленных букв
    trans = {
        "К": "K", "С": "C", "А": "A", "В": "B", "М": "M",
        "Н": "H", "Р": "P", "Т": "T", "О": "O", "Е": "E", "Х": "X"
    }
    s = "".join(trans.get(ch, ch) for ch in s)
    return s


_ASSISTANT_PRINCIPALS = {
    # Lydia Cherkasova
    "0M": "Лидия Черкасова | Lydia Cherkasova",
    "0T": "Лидия Черкасова | Lydia Cherkasova",
    "1A": "Лидия Черкасова | Lydia Cherkasova",
    "1G": "Лидия Черкасова | Lydia Cherkasova",
    "1P": "Лидия Черкасова | Lydia Cherkasova",   # Р -> P
    "2A": "Лидия Черкасова | Lydia Cherkasova",
    "2D": "Лидия Черкасова | Lydia Cherkasova",
    "2P": "Лидия Черкасова | Lydia Cherkasova",
    "3A": "Лидия Черкасова | Lydia Cherkasova",
    "3M": "Лидия Черкасова | Lydia Cherkasova",
    "3H": "Лидия Черкасова | Lydia Cherkasova",   # Н -> H

    # Alexander Belov
    "4A": "Александр Белов | Alexander Belov",
    "4Б": "Александр Белов | Alexander Belov",
    "4K1/4C1": "Александр Белов | Alexander Belov",
    "4K2/4C2": "Александр Белов | Alexander Belov",
    "5A": "Александр Белов | Alexander Belov",
    "5K1/5C1": "Александр Белов | Alexander Belov",
    "5K2/5C2": "Александр Белов | Alexander Belov",
    "6A": "Александр Белов | Alexander Belov",
    "6K1/6C1": "Александр Белов | Alexander Belov",
    "6K2/6C2": "Александр Белов | Alexander Belov",

    # Olga Tkachuk
    "7A": "Ольга Ткачук | Olga Tkachuk",
    "7K/7C": "Ольга Ткачук | Olga Tkachuk",
    "8A": "Ольга Ткачук | Olga Tkachuk",
    "8K/8C": "Ольга Ткачук | Olga Tkachuk",
    "9A": "Ольга Ткачук | Olga Tkachuk",
    "9K/9C": "Ольга Ткачук | Olga Tkachuk",
    "10A": "Ольга Ткачук | Olga Tkachuk",
    "10K/10C": "Ольга Ткачук | Olga Tkachuk",
    "11A": "Ольга Ткачук | Olga Tkachuk",
    "11K/11C": "Ольга Ткачук | Olga Tkachuk",
}

def get_assistant_principal_for_class(level: str) -> str:
    """
    Возвращает строку «Имя Фамилия | Name Surname» для нужного класса.
    Если класс не найден — вернёт пустую строку.
    """
    key = _normalize_class(level)
    return _ASSISTANT_PRINCIPALS.get(key, "")

# --- TUTOR BLOCK -------------------------------------------------------------
from openpyxl import load_workbook

def get_tutor_info_from_xlsx(xlsx_path: str, class_code: str, debug: bool = False) -> tuple[str, str]:
    """
    Возвращает (tutor_text, tutor_name) для указанного класса из листа,
    название которого содержит 'Тьютор' или 'Tutor' (без учёта регистра).
    Переносы строк сохраняем.
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        if debug:
            print(f"[TUTOR][ERR] Не удалось открыть книгу: {xlsx_path}. Ошибка: {e}")
        return "", ""

    # ищем лист по подстроке 'тьютор'/'tutor'
    ws = None
    for name in wb.sheetnames:
        low = name.lower().strip()
        if ("тьютор" in low) or ("tutor" in low):
            ws = wb[name]
            if debug:
                print(f"[TUTOR] Используем лист: '{name}' (max_row={ws.max_row}, max_col={ws.max_column})")
            break
    if ws is None:
        if debug:
            print(f"[TUTOR] Лист с 'Тьютор'/'Tutor' не найден. Доступные листы: {wb.sheetnames}")
        return "", ""

    # 1) Находим колонку класса по шапке (ищем в 1..3 строках)
    header_rows = (1, 2, 3)
    class_col_idx = None
    class_code_norm = str(class_code).strip()

    for r in header_rows:
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip() == class_code_norm:
                class_col_idx = c
                break
        if class_col_idx:
            break
    if debug:
        print(f"[TUTOR] class_code='{class_code_norm}', найденный столбец: {class_col_idx}")
    if not class_col_idx:
        return "", ""

    # 2) Собираем текст тьютора до «дыры» 3 пустых строки подряд (сохраняем \n)
    lines, empty_streak = [], 0
    start_row = max(header_rows) + 1
    for r in range(start_row, ws.max_row + 1):
        left = ws.cell(row=r, column=1).value
        left_s = (str(left).strip().lower() if left is not None else "")

        # если слева пошли заголовки таблицы — стоп
        if any(k in left_s for k in ("дескриптор", "descriptor",
                                     "критерии", "assessment",
                                     "имя", "фамилия")):
            break

        cell_val = ws.cell(row=r, column=class_col_idx).value
        text_raw = "" if cell_val is None else str(cell_val)
        text = text_raw.strip()

        # если в столбце класса встретилось слово "комментарий" — это шапка персональных комментов → стоп
        if text and ("коммент" in text.lower()):
            break

        if text == "":
            empty_streak += 1
            if empty_streak >= 3:
                break
        else:
            empty_streak = 0
            lines.append(text)

    tutor_text = "\n".join(lines).strip()  # <— ключевое отличие: сохраняем переносы строк

    # 3) Имя тьютора: строка, где в первом столбце есть 'Тьютор'
    tutor_name = ""
    for r in range(1, ws.max_row + 1):
        left_label = ws.cell(row=r, column=1).value
        if left_label and "Тьютор" in str(left_label):
            val = ws.cell(row=r, column=class_col_idx).value
            tutor_name = (str(val).strip() if val else "")
            break

    if debug:
        preview = (tutor_text[:80] + "…") if len(tutor_text) > 80 else tutor_text
        print(f"[TUTOR] Собран текст ({len(tutor_text)} симв.): '{preview}'")
        print(f"[TUTOR] Имя тьютора: '{tutor_name}'")
        print("[TUTOR] --- конец отладки ---")

    return tutor_text, tutor_name
