from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.utils import timezone

from validation.services import WorkbookReadError, parse_subject_sheet

from .models import AssessmentCriterion, ClassWorkbook, Student, StudentAssessment, SubjectSheet


@dataclass
class ParseWorkbookResult:
    workbook_id: int
    class_code: str
    sheets_total: int
    sheets_imported: int
    tutor_sheets: int
    sheets_skipped: int
    criteria_created: int
    students_created: int
    assessments_created: int


def _normalize_name(first_name: str, last_name: str) -> str:
    raw = f"{last_name} {first_name}".strip().lower()
    raw = re.sub(r"\s+", " ", raw)
    return re.sub(r"[^a-zа-яё0-9]+", "_", raw).strip("_")


def _is_tutor_sheet(sheet_name: str) -> bool:
    title = (sheet_name or "").strip().lower()
    return "тьютор" in title or "tutor" in title


def _parse_module_number(raw_value: object) -> int:
    if raw_value is None:
        return 0
    text = str(raw_value).strip()
    if not text:
        return 0
    try:
        return int(float(text))
    except (TypeError, ValueError):
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else 0


def _as_decimal(raw_value: object) -> Decimal | None:
    if raw_value is None:
        return None
    text = str(raw_value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def import_class_workbook(
    *,
    workbook_path: str,
    class_code: str,
    source_url: str,
    period: str,
    fetched_at: datetime | None = None,
) -> ParseWorkbookResult:
    """Parse all workbook sheets and write data into DATA-101 tables."""
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from zipfile import BadZipFile

    try:
        wb = load_workbook(workbook_path, data_only=True)
    except (FileNotFoundError, InvalidFileException, BadZipFile, OSError) as exc:
        raise WorkbookReadError(f"Cannot read workbook: {workbook_path}") from exc

    fetched_ts = fetched_at or timezone.now()

    with transaction.atomic():
        class_workbook = ClassWorkbook.objects.create(
            class_code=class_code,
            source_url=source_url,
            period=period,
            fetched_at=fetched_ts,
        )

        sheets_imported = 0
        tutor_sheets = 0
        sheets_skipped = 0
        criteria_created = 0
        students_created = 0
        assessments_created = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parsed = parse_subject_sheet(ws)
            is_tutor = _is_tutor_sheet(sheet_name)

            teacher_name = parsed["metadata"].get("teacher", "")
            module_number = _parse_module_number(parsed["metadata"].get("module"))
            descriptor_text = parsed["metadata"].get("descriptor", "")

            subject_sheet = SubjectSheet.objects.create(
                workbook=class_workbook,
                sheet_name=sheet_name,
                subject_name=sheet_name,
                teacher_name=teacher_name,
                module_number=module_number,
                descriptor_text=descriptor_text,
                is_tutor=is_tutor,
            )
            sheets_imported += 1
            if is_tutor:
                tutor_sheets += 1
                continue

            criteria_row = parsed.get("criteria_row")
            if not criteria_row:
                sheets_skipped += 1
                continue

            criterion_by_col: dict[int, AssessmentCriterion] = {}

            for col in parsed.get("criteria_cols", []):
                title = str(ws.cell(row=criteria_row, column=col).value or "").strip()
                criterion = AssessmentCriterion.objects.create(
                    subject_sheet=subject_sheet,
                    column_index=col,
                    criterion_text=title,
                    criterion_type=AssessmentCriterion.CriterionType.CRITERION,
                )
                criterion_by_col[col] = criterion
                criteria_created += 1

            for col in parsed.get("test_cols", []):
                title = str(ws.cell(row=criteria_row, column=col).value or "").strip()
                criterion = AssessmentCriterion.objects.create(
                    subject_sheet=subject_sheet,
                    column_index=col,
                    criterion_text=title,
                    criterion_type=AssessmentCriterion.CriterionType.TEST,
                )
                criterion_by_col[col] = criterion
                criteria_created += 1

            if parsed.get("comment_col"):
                col = parsed["comment_col"]
                title = str(ws.cell(row=criteria_row, column=col).value or "Комментарий").strip()
                criterion = AssessmentCriterion.objects.create(
                    subject_sheet=subject_sheet,
                    column_index=col,
                    criterion_text=title,
                    criterion_type=AssessmentCriterion.CriterionType.COMMENT,
                )
                criterion_by_col[col] = criterion
                criteria_created += 1

            if parsed.get("retake_col"):
                col = parsed["retake_col"]
                title = str(ws.cell(row=criteria_row, column=col).value or "Пересдача").strip()
                criterion = AssessmentCriterion.objects.create(
                    subject_sheet=subject_sheet,
                    column_index=col,
                    criterion_text=title,
                    criterion_type=AssessmentCriterion.CriterionType.RETAKE,
                )
                criterion_by_col[col] = criterion
                criteria_created += 1

            for student_info in parsed.get("students", []):
                first_name = str(student_info.get("first") or "").strip()
                last_name = str(student_info.get("last") or "").strip()
                normalized = _normalize_name(first_name, last_name)

                student, created = Student.objects.get_or_create(
                    class_code=class_code,
                    full_name_normalized=normalized,
                    defaults={
                        "first_name": first_name,
                        "last_name": last_name,
                    },
                )
                if created:
                    students_created += 1

                row_idx = int(student_info["row"])

                for col in parsed.get("criteria_cols", []):
                    raw = ws.cell(row=row_idx, column=col).value
                    StudentAssessment.objects.create(
                        subject_sheet=subject_sheet,
                        student=student,
                        criterion=criterion_by_col.get(col),
                        raw_value=str(raw or "").strip(),
                    )
                    assessments_created += 1

                for col in parsed.get("test_cols", []):
                    raw = ws.cell(row=row_idx, column=col).value
                    StudentAssessment.objects.create(
                        subject_sheet=subject_sheet,
                        student=student,
                        criterion=criterion_by_col.get(col),
                        raw_value=str(raw or "").strip(),
                        numeric_score=_as_decimal(raw),
                    )
                    assessments_created += 1

                if parsed.get("comment_col"):
                    col = parsed["comment_col"]
                    raw = ws.cell(row=row_idx, column=col).value
                    StudentAssessment.objects.create(
                        subject_sheet=subject_sheet,
                        student=student,
                        criterion=criterion_by_col.get(col),
                        raw_value=str(raw or "").strip(),
                        comment_text=str(raw or "").strip(),
                    )
                    assessments_created += 1

                if parsed.get("retake_col"):
                    col = parsed["retake_col"]
                    raw = ws.cell(row=row_idx, column=col).value
                    StudentAssessment.objects.create(
                        subject_sheet=subject_sheet,
                        student=student,
                        criterion=criterion_by_col.get(col),
                        raw_value=str(raw or "").strip(),
                        retake_flag=bool(str(raw or "").strip()),
                    )
                    assessments_created += 1

    return ParseWorkbookResult(
        workbook_id=class_workbook.id,
        class_code=class_code,
        sheets_total=len(wb.sheetnames),
        sheets_imported=sheets_imported,
        tutor_sheets=tutor_sheets,
        sheets_skipped=sheets_skipped,
        criteria_created=criteria_created,
        students_created=students_created,
        assessments_created=assessments_created,
    )