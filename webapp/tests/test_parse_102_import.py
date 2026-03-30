import json
import tempfile
from pathlib import Path

from django.core.management import call_command
from django.test import TestCase
from openpyxl import Workbook

from webapp.models import AssessmentCriterion, ClassWorkbook, StudentAssessment, SubjectSheet
from webapp.parsing import import_class_workbook


def _build_4a_workbook(path: Path) -> None:
    wb = Workbook()

    ws_math = wb.active
    ws_math.title = "Math"
    ws_math["C1"] = "4A"
    ws_math["C2"] = "Ms. Math"
    ws_math["C3"] = "Module 1"
    ws_math["C4"] = "Math descriptor"
    ws_math.cell(row=5, column=1, value="Name")
    ws_math.cell(row=5, column=2, value="Surname")
    ws_math.cell(row=5, column=3, value="Критерии оценивания | Assessment criteria")
    ws_math.cell(row=5, column=4, value="Решает задачи")
    ws_math.cell(row=5, column=5, value="Quiz 1")
    ws_math.cell(row=5, column=6, value="Comment")
    ws_math.cell(row=5, column=7, value="Retake")
    ws_math.cell(row=6, column=1, value="Ivan")
    ws_math.cell(row=6, column=2, value="Ivanov")
    ws_math.cell(row=6, column=4, value="соответствует ожиданиям")
    ws_math.cell(row=6, column=5, value="89")
    ws_math.cell(row=6, column=6, value="Good")
    ws_math.cell(row=6, column=7, value="")

    ws_eng = wb.create_sheet("English")
    ws_eng["C1"] = "4A"
    ws_eng["C2"] = "Ms. English"
    ws_eng["C3"] = 2
    ws_eng["C4"] = "English descriptor"
    ws_eng.cell(row=5, column=1, value="Name")
    ws_eng.cell(row=5, column=2, value="Surname")
    ws_eng.cell(row=5, column=3, value="Критерии оценивания | Assessment criteria")
    ws_eng.cell(row=5, column=4, value="Reads text")
    ws_eng.cell(row=5, column=5, value="Test 1")
    ws_eng.cell(row=5, column=6, value="Комментарий")
    ws_eng.cell(row=5, column=7, value="Пересдача")
    ws_eng.cell(row=6, column=1, value="Ivan")
    ws_eng.cell(row=6, column=2, value="Ivanov")
    ws_eng.cell(row=6, column=4, value="начальный уровень")
    ws_eng.cell(row=6, column=5, value=74)
    ws_eng.cell(row=6, column=6, value="Needs practice")
    ws_eng.cell(row=6, column=7, value="yes")

    ws_tutor = wb.create_sheet("Тьютор | Tutor")
    ws_tutor["C1"] = "4A"
    ws_tutor["C2"] = "Tutor Name"
    ws_tutor["C3"] = "1"
    ws_tutor["C4"] = "Tutor descriptor"

    wb.save(path)


class Parse102WorkbookImportTests(TestCase):
    def test_orchestrator_imports_all_sheets_and_marks_tutor(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "journal_4A.xlsx"
            _build_4a_workbook(workbook_path)

            result = import_class_workbook(
                workbook_path=str(workbook_path),
                class_code="4A",
                source_url="https://docs.google.com/spreadsheets/d/workbook-4a",
                period="module-1",
            )

        self.assertEqual(result.sheets_total, 3)
        self.assertEqual(result.sheets_imported, 3)
        self.assertEqual(result.tutor_sheets, 1)
        self.assertGreaterEqual(result.criteria_created, 8)
        self.assertGreaterEqual(result.assessments_created, 8)

        workbook = ClassWorkbook.objects.get(id=result.workbook_id)
        self.assertEqual(workbook.class_code, "4A")

        sheets = SubjectSheet.objects.filter(workbook=workbook)
        self.assertEqual(sheets.count(), 3)

        tutor_sheet = sheets.get(sheet_name="Тьютор | Tutor")
        self.assertTrue(tutor_sheet.is_tutor)

        for subject_name in ["Math", "English"]:
            subject = sheets.get(sheet_name=subject_name)
            self.assertFalse(subject.is_tutor)
            self.assertNotEqual(subject.teacher_name, "")
            self.assertGreater(subject.module_number, 0)
            self.assertNotEqual(subject.descriptor_text, "")
            self.assertTrue(AssessmentCriterion.objects.filter(subject_sheet=subject).exists())
            self.assertTrue(StudentAssessment.objects.filter(subject_sheet=subject).exists())

    def test_management_command_outputs_parse_statistics(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "journal_4A.xlsx"
            _build_4a_workbook(workbook_path)

            from io import StringIO

            out = StringIO()
            call_command(
                "import_class_workbook",
                "--workbook-path",
                str(workbook_path),
                "--class-code",
                "4A",
                "--source-url",
                "https://docs.google.com/spreadsheets/d/workbook-4a",
                "--period",
                "module-1",
                stdout=out,
            )

        payload = json.loads(out.getvalue())
        self.assertEqual(payload["class_code"], "4A")
        self.assertEqual(payload["summary"]["sheets_total"], 3)
        self.assertEqual(payload["summary"]["tutor_sheets"], 1)