import logging
import re
from zipfile import BadZipFile
from dataclasses import asdict, dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


from .rules import (
    ALLOWED_DESCRIPTOR_VALUES,
    COMMENT_REQUIRED_IF_LOW_SCORE,
    LOW_SCORE_THRESHOLD,
    REQUIRED_HEADER_KEYS,
    RETAKE_REQUIRED_IF_LOW_SCORE,
    TEST_SCORE_MAX,
    TEST_SCORE_MIN,
    VALID_RETAKE_VALUES,
)

_CRITERIA_ANCHOR_RU = "критерии оценивания"
_CRITERIA_ANCHOR_EN = "assessment criteria"
@dataclass
class ValidationIssue:
    code: str
    severity: str  # critical | warning | info
    sheet: str
    row: int | None
    student: str | None
    field: str | None
    message: str
    class_code: str | None = None
    subject_name: str | None = None
    teacher_name: str | None = None
    module_number: int | None = None
    column_type: str | None = None
    issue_group: str | None = None
    missing_count: int = 1


class WorkbookReadError(ValueError):
    """Raised when a workbook cannot be read for validation."""


logger = logging.getLogger(__name__)
def _classify_sheet_type(sheet_name: str) -> str:
    normalized_name = (sheet_name or "").strip().lower()
    if "тьютор" in normalized_name or "tutor" in normalized_name:
        return "tutor"
    if "служеб" in normalized_name or "service" in normalized_name:
        return "service"
    return "subject"
def _is_empty(v: Any) -> bool:
    return v is None or str(v).strip() == ""

def _normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()

def _get_real_data_bounds(ws) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for row_num in range(1, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            if not _is_empty(ws.cell(row=row_num, column=col_num).value):
                max_row = max(max_row, row_num)
                max_col = max(max_col, col_num)
    return max_row, max_col

def _find_criteria_header_row(ws, max_row: int, max_col: int) -> int | None:
    for row_num in range(1, max_row + 1):
        anchor_row_text = " | ".join(
            _normalize_text(ws.cell(row=row_num, column=col).value)
            for col in range(1, max_col + 1)
        )
        if _CRITERIA_ANCHOR_RU in anchor_row_text and _CRITERIA_ANCHOR_EN in anchor_row_text:
            return row_num

    # Fallback for legacy sheets without explicit RU/EN anchor
    for row_num in range(1, max_row + 1):
        row_headers = [
            _normalize_text(ws.cell(row=row_num, column=col).value)
            for col in range(1, max_col + 1)
        ]
        has_name_columns = any(h in ("имя", "name", "first name") for h in row_headers) and any(
            "фам" in h or h in ("surname", "last name") for h in row_headers
        )
        has_assessment_columns = any(
            any(k in h for k in ("критер", "тест", "quiz", "comment", "коммент", "retake", "пересда"))
            for h in row_headers
        )
        if has_name_columns and has_assessment_columns:
            return row_num

    return None


def _detect_student_header_row(ws, criteria_row: int, max_row: int, max_col: int) -> int:
    for row_num in range(criteria_row, max_row + 1):
        row_headers = [
            _normalize_text(ws.cell(row=row_num, column=col).value)
            for col in range(1, max_col + 1)
        ]
        has_first = any(h in ("имя", "name", "first name") for h in row_headers)
        has_last = any("фам" in h or h in ("surname", "last name") for h in row_headers)
        if has_first and has_last:
            return row_num
    return criteria_row

def _detect_student_name_columns(ws, header_row: int, max_col: int) -> tuple[int, int]:
    first_name_col = 1
    last_name_col = 2

    for col in range(1, max_col + 1):
        hdr = _normalize_text(ws.cell(row=header_row, column=col).value)
        if any(token in hdr for token in ("имя", "first", "name")):
            first_name_col = col
        if any(token in hdr for token in ("фам", "last", "surname")):
            last_name_col = col

    return first_name_col, last_name_col


def parse_subject_sheet(ws) -> dict:
    """Parse one subject sheet with dynamic anchors/columns/students range."""
    metadata = {
        "class": str(ws.cell(row=1, column=3).value or "").strip(),
        "teacher": str(ws.cell(row=2, column=3).value or "").strip(),
        "module": str(ws.cell(row=3, column=3).value or "").strip(),
        "descriptor": str(ws.cell(row=4, column=3).value or "").strip(),
    }

    max_row, max_col = _get_real_data_bounds(ws)
    if max_row == 0 or max_col == 0:
        return {
            "metadata": metadata,
            "criteria_row": None,
            "student_header_row": None,
            "student_start_row": None,
            "first_name_col": 1,
            "last_name_col": 2,
            "criteria_cols": [],
            "test_cols": [],
            "comment_col": None,
            "retake_col": None,
            "students": [],
            "column_types": {},
        }
    criteria_row = _find_criteria_header_row(ws, max_row=max_row, max_col=max_col)
    if criteria_row is None:
        return {
            "metadata": metadata,
            "criteria_row": None,
            "student_header_row": None,
            "student_start_row": None,
            "first_name_col": 1,
            "last_name_col": 2,
            "criteria_cols": [],
            "test_cols": [],
            "comment_col": None,
            "retake_col": None,
            "column_types": {},
            "students": [],
        }

    student_header_row = _detect_student_header_row(ws, criteria_row=criteria_row, max_row=max_row, max_col=max_col)
    student_start_row = student_header_row + 1
    first_name_col, last_name_col = _detect_student_name_columns(ws, student_header_row, max_col=max_col)

    comment_col = None
    retake_col = None
    test_cols: list[int] = []
    criteria_cols: list[int] = []
    column_types: dict[int, str] = {}

    for col in range(1, max_col + 1):
        criteria_hdr = _normalize_text(ws.cell(row=criteria_row, column=col).value)
        student_hdr = _normalize_text(ws.cell(row=student_header_row, column=col).value)
        hdr = criteria_hdr or student_hdr
        if not hdr:
            continue

        if col in (first_name_col, last_name_col):
            continue
        if _CRITERIA_ANCHOR_RU in criteria_hdr or _CRITERIA_ANCHOR_EN in criteria_hdr:
            continue

        if any(k in hdr for k in ("коммент", "comment")):
            comment_col = col
            column_types[col] = "comment"
            continue
        if any(k in hdr for k in ("пересда", "retake", "resit", "make-up")):
            retake_col = col
            column_types[col] = "retake"
            continue
        if any(k in hdr for k in ("квиз", "тест", "quiz", "test")):
            test_cols.append(col)
            column_types[col] = "test"
            continue

        criteria_cols.append(col)
        column_types[col] = "criterion"

    students = []
    started = False
    for row in range(student_start_row, max_row + 1):
        first = ws.cell(row=row, column=first_name_col).value
        last = ws.cell(row=row, column=last_name_col).value

        if _is_empty(first) and _is_empty(last):
            if started:
                break
            continue

        started = True
        student_name = f"{first or ''} {last or ''}".strip()

        criteria_values = {
            str(ws.cell(row=criteria_row, column=col).value or "").strip(): ws.cell(row=row, column=col).value
            for col in criteria_cols
        }
        test_values = {
            str(ws.cell(row=criteria_row, column=col).value or "").strip(): ws.cell(row=row, column=col).value
            for col in test_cols
        }

        students.append(
            {
                "row": row,
                "first": first,
                "last": last,
                "student": student_name,
                "criteria_values": criteria_values,
                "test_values": test_values,
                "comment": ws.cell(row=row, column=comment_col).value if comment_col else None,
                "retake": ws.cell(row=row, column=retake_col).value if retake_col else None,
            }
        )

    return {
        "metadata": metadata,
        "criteria_row": criteria_row,
        "student_header_row": student_header_row,
        "student_start_row": student_start_row,
        "first_name_col": first_name_col,
        "last_name_col": last_name_col,
        "criteria_cols": criteria_cols,
        "test_cols": test_cols,
        "comment_col": comment_col,
        "retake_col": retake_col,
        "column_types": column_types,
        "students": students,
    }

def validate_workbook(path: str) -> dict:
    wb = None
    try:
        wb = load_workbook(path, data_only=True)
    except (FileNotFoundError, InvalidFileException, BadZipFile, OSError) as exc:
        logger.error(
            "Failed to read workbook for validation: %s (%s)",
            path,
            exc.__class__.__name__,
        )
        raise WorkbookReadError(f"Cannot read workbook: {path}") from exc

    issues: list[ValidationIssue] = []
    sheets_skipped = 0
    sheets_validated = 0
    students_total = 0
    issues_by_code: dict[str, int] = {}
    sheet_events: list[dict[str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_type = _classify_sheet_type(sheet_name)
        sheet_events.append({"event": "sheet_detected", "sheet_name": sheet_name, "sheet_type": sheet_type})
        if sheet_type in {"tutor", "service"}:
            sheets_skipped += 1
            sheet_events.append({"event": "sheet_skipped", "sheet_name": sheet_name, "sheet_type": sheet_type})
            logger.info(
                "Skipping sheet '%s' with type '%s' in validation profile",
                sheet_name,
                sheet_type,
            )
            continue
        logger.debug(
            "Validating sheet '%s' with type '%s' using subject profile",
            sheet_name,
            sheet_type,
        )
        sheets_validated += 1
        parsed = parse_subject_sheet(ws)
        students_total += len(parsed.get("students", []))
        sheet_issues = validate_sheet(ws, sheet_name, parsed=parsed)
        issues.extend(sheet_issues)
        sheet_events.append({"event": "sheet_validated", "sheet_name": sheet_name, "sheet_type": sheet_type})

    for issue in issues:
        issues_by_code[issue.code] = issues_by_code.get(issue.code, 0) + 1

    summary = {
        "total": len(issues),
        "critical": sum(1 for i in issues if i.severity == "critical"),
        "warning": sum(1 for i in issues if i.severity == "warning"),
        "info": sum(1 for i in issues if i.severity == "info"),
        "sheets_total": len(wb.sheetnames),
        "sheets_validated": sheets_validated,
        "sheets_skipped": sheets_skipped,
        "students_total": students_total,
        "issues_by_code": issues_by_code,
    }

    try:
        return {
            "summary": summary,
            "issues": [asdict(i) for i in issues],
            "sheet_events": sheet_events,
        }
    finally:
        if wb is not None:
            wb.close()


def validate_sheet(ws, sheet_name: str, parsed: dict | None = None) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    parsed = parsed or parse_subject_sheet(ws)
    metadata = parsed.get("metadata", {})
    for key in REQUIRED_HEADER_KEYS:
        metadata.setdefault(key, "")

    class_meta = str(metadata.get("class", "")).strip()
    teacher_meta = str(metadata.get("teacher", "")).strip()
    module_meta = str(metadata.get("module", "")).strip()
    module_number = int(float(module_meta)) if _is_numeric(module_meta) else None
    column_types = parsed.get("column_types", {})

    def make_issue(
            *,
            code: str,
            severity: str,
            row: int | None,
            student: str | None,
            field: str | None,
            message: str,
            issue_group: str | None = None,
            missing_count: int = 1,
    ) -> ValidationIssue:
        column_type = None
        if field and field.startswith("col_"):
            try:
                col_number = int(field.replace("col_", "", 1))
            except ValueError:
                col_number = None
            if col_number is not None:
                column_type = column_types.get(col_number)

        return ValidationIssue(
            code=code,
            severity=severity,
            sheet=sheet_name,
            row=row,
            student=student,
            field=field,
            message=message,
            class_code=class_meta or None,
            subject_name=sheet_name or None,
            teacher_name=teacher_meta or None,
            module_number=module_number,
            column_type=column_type,
            issue_group=issue_group,
            missing_count=missing_count,
        )

    if _is_empty(class_meta):
        issues.append(make_issue(
            code="MISSING_CLASS_META",
            severity="critical",
            row=1,
            student=None,
            field="meta_class",
            message="Не заполнено поле «Класс | Grade» (C1). Заполните класс в ячейке C1.",
        ))

    if _is_empty(teacher_meta):
        issues.append(make_issue(
            code="MISSING_TEACHER_META",
            severity="critical",
            row=2,
            student=None,
            field="meta_teacher",
            message="Не заполнено поле «Учитель | Teacher» (C2). Укажите ФИО учителя в ячейке C2.",
        ))

    if not _is_numeric(module_meta):
        issues.append(make_issue(
            code="INVALID_MODULE_META",
            severity="warning",
            row=3,
            student=None,
            field="meta_module",
            message="Поле «Учебный модуль | Module» (C3) должно быть числом. Укажите номер модуля цифрой (например, 1, 2, 3).",
        ))

    descriptor_meta = str(metadata.get("descriptor", "")).strip()
    if _is_empty(descriptor_meta):
        issues.append(make_issue(
            code="DESCRIPTOR_EMPTY",
            severity="critical",
            row=4,
            student=None,
            field="meta_descriptor",
            message="Не заполнено поле «Дескриптор | Descriptor» (C4). Заполните описание дескриптора в ячейке C4.",
            issue_group="descriptor",
        ))

    criteria_cols = parsed["criteria_cols"]
    test_cols = parsed["test_cols"]
    comment_col = parsed["comment_col"]
    retake_col = parsed["retake_col"]
    for student_info in parsed["students"]:
        row = student_info["row"]
        student_name = student_info["student"]

        for col in criteria_cols:
            v = ws.cell(row=row, column=col).value
            if _is_empty(v):
                issues.append(make_issue(
                    code="CRITERIA_HEADERS_EMPTY",
                    severity="critical",
                    row=row,
                    student=student_name,
                    field=f"col_{col}",
                    message="Не заполнен критерий оценивания",
                    issue_group="criteria",
                ))
            else:
                sv = str(v).strip()
                if _is_numeric(sv):
                    issues.append(make_issue(
                        code="CRITERION_EXPECTS_LEVEL",
                        severity="warning",
                        row=row,
                        student=student_name,
                        field=f"col_{col}",
                        message=f"В колонке критерия ожидается уровень, получено число: {sv}",
                    ))
                elif sv not in ALLOWED_DESCRIPTOR_VALUES:
                    issues.append(make_issue(
                        code="INVALID_CRITERION_VALUE",
                        severity="warning",
                        row=row,
                        student=student_name,
                        field=f"col_{col}",
                        message=f"Недопустимое значение критерия: '{sv}'",
                    ))

        low_score_found = False
        low_score_context: tuple[int, float] | None = None
        for col in test_cols:
            raw = ws.cell(row=row, column=col).value
            if _is_empty(raw):
                issues.append(make_issue(
                    code="GRADE_EMPTY",
                    severity="critical",
                    row=row,
                    student=student_name,
                    field=f"col_{col}",
                    message="Не заполнена оценка (тестовый балл)",
                    issue_group="grades",
                ))
                continue
            try:
                score = float(raw)
                if score < TEST_SCORE_MIN or score > TEST_SCORE_MAX:
                    issues.append(make_issue(
                        code="TEST_SCORE_OUT_OF_RANGE",
                        severity="warning",
                        row=row,
                        student=student_name,
                        field=f"col_{col}",
                        message=f"Баллы вне диапазона {TEST_SCORE_MIN}-{TEST_SCORE_MAX}: {score}",
                    ))
                if score < LOW_SCORE_THRESHOLD:
                    low_score_found = True
                    if low_score_context is None:
                        low_score_context = (col, score)
            except Exception:
                issues.append(make_issue(
                    code="TEST_SCORE_NOT_NUMERIC",
                    severity="warning",
                    row=row,
                    student=student_name,
                    field=f"col_{col}",
                    message=f"Тестовый балл не число: {raw}",
                ))
        if retake_col and column_types.get(retake_col) == "retake":
            raw_retake = ws.cell(row=row, column=retake_col).value
            if not _is_empty(raw_retake):
                normalized_retake = _normalize_text(raw_retake)
                if normalized_retake not in VALID_RETAKE_VALUES:
                    issues.append(make_issue(
                        code="INVALID_RETAKE_VALUE",
                        severity="warning",
                        row=row,
                        student=student_name,
                        field=f"col_{retake_col}",
                        message=f"Недопустимое значение пересдачи: {raw_retake}",
                    ))
        if low_score_found:
            low_score_message_suffix = ""
            if low_score_context:
                low_col, low_score = low_score_context
                low_score_message_suffix = (
                    f" (тестовая колонка: col_{low_col}, значение: {low_score:g})"
                )
            if COMMENT_REQUIRED_IF_LOW_SCORE and comment_col:
                c = ws.cell(row=row, column=comment_col).value
                if _is_empty(c):
                    issues.append(make_issue(
                        code="COMMENT_REQUIRED",
                        severity="critical",
                        row=row,
                        student=student_name,
                        field=f"col_{comment_col}",
                        message=(
                            "Нужен комментарий при низком тестовом балле"
                            f"{low_score_message_suffix}"
                        ),
                        issue_group="grades",
                    ))
            if RETAKE_REQUIRED_IF_LOW_SCORE and retake_col:
                r = ws.cell(row=row, column=retake_col).value
                if _is_empty(r):
                    issues.append(make_issue(
                        code="RETAKE_REQUIRED",
                        severity="critical",
                        row=row,
                        student=student_name,
                        field=f"col_{retake_col}",
                        message=(
                            "Нужно указать пересдачу при низком тестовом балле"
                            f"{low_score_message_suffix}"
                        ),
                        issue_group="grades",
                    ))


    return issues

def _is_numeric(value: Any) -> bool:
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False