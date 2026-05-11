import tempfile
from datetime import timedelta
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth.models import Permission, User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from jobs.models import JobRun
from journal_links.models import ClassSheetLink, DescriptorCriteriaCheckSchedule, DescriptorCriteriaReportTarget
from validation.descriptor_criteria_fill import (
    JOB_TYPE,
    check_workbook_descriptor_criteria,
    enqueue_descriptor_criteria_fill_check_job,
    run_descriptor_criteria_fill_check_job,
)


def _build_descriptor_criteria_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Math"
    ws.cell(row=1, column=2, value="Класс | Grade")
    ws.cell(row=1, column=3, value="7A")
    ws.cell(row=2, column=2, value="Учитель | Teacher")
    ws.cell(row=2, column=3, value="Teacher A")
    ws.cell(row=3, column=2, value="Модуль | Module")
    ws.cell(row=3, column=3, value="2")
    ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
    ws.cell(row=4, column=3, value="Module descriptor")
    ws.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
    ws.cell(row=6, column=3, value="Criterion 1")
    ws.cell(row=6, column=4, value="")
    ws.cell(row=6, column=5, value="Комментарий")
    ws.cell(row=8, column=1, value="Student")
    ws.cell(row=8, column=3, value="")

    ws_tutor = wb.create_sheet("Tutor support")
    ws_tutor.cell(row=4, column=2, value="Дескриптор | Descriptor")
    ws_tutor.cell(row=4, column=3, value="")
    ws_tutor.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
    ws_tutor.cell(row=6, column=3, value="")

    ws_service = wb.create_sheet("Service data")
    ws_service.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")

    wb.save(path)


def _build_grade_completeness_workbook(path: Path, *, missing_cells: set[tuple[int, int]] | None = None) -> None:
    missing_cells = missing_cells or set()
    wb = Workbook()
    ws = wb.active
    ws.title = "Math"
    ws.cell(row=1, column=2, value="Класс | Grade")
    ws.cell(row=1, column=3, value="7A")
    ws.cell(row=2, column=2, value="Учитель | Teacher")
    ws.cell(row=2, column=3, value="Teacher A")
    ws.cell(row=3, column=2, value="Модуль | Module")
    ws.cell(row=3, column=3, value="1")
    ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
    ws.cell(row=4, column=3, value="Descriptor")
    ws.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
    ws.cell(row=6, column=3, value="Criterion 1")
    ws.cell(row=6, column=4, value="Criterion 2")
    ws.cell(row=6, column=5, value="Criterion 3")
    ws.cell(row=6, column=6, value="")
    ws.cell(row=6, column=7, value="Комментарий")
    ws.cell(row=8, column=1, value="Имя")
    ws.cell(row=8, column=2, value="Фамилия")
    ws.cell(row=9, column=1, value="Ada")
    ws.cell(row=9, column=2, value="Lovelace")
    ws.cell(row=10, column=1, value="Grace")
    ws.cell(row=10, column=2, value="Hopper")
    for row_num in (9, 10):
        for col_num in (3, 4, 5):
            if (row_num, col_num) not in missing_cells:
                ws.cell(row=row_num, column=col_num, value=5)
    wb.save(path)


class DescriptorCriteriaFillWorkbookTests(TestCase):
    def test_checks_subject_sheets_and_skips_tutor_service(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "descriptor_criteria.xlsx"
            _build_descriptor_criteria_workbook(workbook_path)

            result = check_workbook_descriptor_criteria(
                str(workbook_path),
                class_code="7A",
                sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            )

        self.assertEqual(result["summary"]["sheets_total"], 3)
        self.assertEqual(result["summary"]["sheets_checked"], 1)
        self.assertEqual(result["summary"]["sheets_skipped"], 2)

        row = result["rows"][0]
        self.assertEqual(row["class_code"], "7A")
        self.assertEqual(row["subject_name"], "Math")
        self.assertEqual(row["teacher_name"], "Teacher A")
        self.assertEqual(row["module_number"], 2)
        self.assertEqual(row["descriptor_status"], "filled")
        self.assertEqual(row["criteria_total"], 2)
        self.assertEqual(row["criteria_filled"], 1)
        self.assertEqual(row["criteria_missing"], 1)
        self.assertEqual(row["overall_status"], "problem")
        self.assertFalse(any(event["sheet_name"] == "Tutor support" and event["event"] == "sheet_checked" for event in result["sheet_events"]))

    def test_student_grade_cells_are_not_counted_as_criteria(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "student_grade_ignored.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Science"
            ws.cell(row=2, column=2, value="Учитель | Teacher")
            ws.cell(row=2, column=3, value="Teacher B")
            ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
            ws.cell(row=4, column=3, value="Descriptor")
            ws.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
            ws.cell(row=6, column=3, value="Criterion 1")
            ws.cell(row=6, column=4, value="Комментарий")
            ws.cell(row=8, column=3, value="")
            wb.save(workbook_path)

            result = check_workbook_descriptor_criteria(
                str(workbook_path),
                class_code="8B",
                sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            )

        self.assertEqual(result["rows"][0]["criteria_total"], 1)
        self.assertEqual(result["rows"][0]["criteria_missing"], 0)
        self.assertEqual(result["rows"][0]["overall_status"], "ok")

    def test_counts_filled_grades_for_student_criteria_intersections(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "grades_filled.xlsx"
            _build_grade_completeness_workbook(workbook_path)

            result = check_workbook_descriptor_criteria(
                str(workbook_path),
                class_code="7A",
                sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            )

        row = result["rows"][0]
        self.assertEqual(row["students_total"], 2)
        self.assertEqual(row["criteria_filled"], 3)
        self.assertEqual(row["criteria_total"], 4)
        self.assertEqual(row["grades_total"], 6)
        self.assertEqual(row["grades_filled"], 6)
        self.assertEqual(row["grades_missing"], 0)
        self.assertEqual(row["grades_ratio"], "6/6")
        self.assertEqual(row["grades_status"], "ok")
        self.assertEqual(row["overall_status"], "problem")

    def test_missing_grade_cells_make_subject_problem(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "grades_missing.xlsx"
            _build_grade_completeness_workbook(workbook_path, missing_cells={(9, 4), (10, 5)})

            result = check_workbook_descriptor_criteria(
                str(workbook_path),
                class_code="7A",
                sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            )

        row = result["rows"][0]
        self.assertEqual(row["grades_total"], 6)
        self.assertEqual(row["grades_filled"], 4)
        self.assertEqual(row["grades_missing"], 2)
        self.assertEqual(row["grades_ratio"], "4/6")
        self.assertEqual(row["grades_status"], "missing")
        self.assertEqual(row["overall_status"], "problem")

    def test_missing_student_header_makes_grades_not_applicable(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "no_students.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Math"
            ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
            ws.cell(row=4, column=3, value="Descriptor")
            ws.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
            ws.cell(row=6, column=3, value="Criterion 1")
            ws.cell(row=6, column=4, value="Комментарий")
            wb.save(workbook_path)

            result = check_workbook_descriptor_criteria(
                str(workbook_path),
                class_code="7A",
                sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            )

        row = result["rows"][0]
        self.assertEqual(row["grades_total"], 0)
        self.assertEqual(row["grades_ratio"], "—")
        self.assertEqual(row["grades_status"], "not_applicable")

    def test_no_filled_criteria_makes_grades_not_applicable(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "no_filled_criteria.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Math"
            ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
            ws.cell(row=4, column=3, value="Descriptor")
            ws.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
            ws.cell(row=6, column=3, value="")
            ws.cell(row=6, column=4, value="Комментарий")
            ws.cell(row=8, column=1, value="Имя")
            ws.cell(row=9, column=1, value="Ada")
            wb.save(workbook_path)

            result = check_workbook_descriptor_criteria(
                str(workbook_path),
                class_code="7A",
                sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            )

        row = result["rows"][0]
        self.assertEqual(row["criteria_filled"], 0)
        self.assertEqual(row["grades_total"], 0)
        self.assertEqual(row["grades_status"], "not_applicable")

    def test_numeric_only_criteria_are_treated_as_missing(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "numeric_criteria.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Math"
            ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
            ws.cell(row=4, column=3, value="Descriptor")
            ws.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
            ws.cell(row=6, column=3, value=1)
            ws.cell(row=6, column=4, value="2")
            ws.cell(row=6, column=5, value="Комментарий")
            ws.cell(row=8, column=1, value="Имя")
            ws.cell(row=9, column=1, value="Ada")
            ws.cell(row=9, column=3, value=5)
            ws.cell(row=9, column=4, value=5)
            wb.save(workbook_path)

            result = check_workbook_descriptor_criteria(
                str(workbook_path),
                class_code="7A",
                sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            )

        row = result["rows"][0]
        self.assertEqual(row["criteria_total"], 2)
        self.assertEqual(row["criteria_filled"], 0)
        self.assertEqual(row["criteria_missing"], 2)
        self.assertEqual(row["grades_total"], 0)
        self.assertEqual(row["grades_status"], "not_applicable")
        self.assertEqual(row["overall_status"], "problem")


class DescriptorCriteriaFillJobTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.link = ClassSheetLink.objects.create(
            class_code="7A",
            subject_name="Math",
            teacher_name="Teacher A",
            google_sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            is_active=True,
        )

    def test_job_creates_result_json_and_logs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "source.xlsx"
            _build_descriptor_criteria_workbook(source_path)
            generated_paths: list[Path] = []

            def _provide_copy(_link):
                path = Path(tmpdir) / f"copy_{len(generated_paths)}.xlsx"
                path.write_bytes(source_path.read_bytes())
                generated_paths.append(path)
                return path

            with patch("validation.descriptor_criteria_fill.fetch_workbook_for_link", side_effect=_provide_copy):
                job_run = run_descriptor_criteria_fill_check_job(class_code="7A")

        self.assertEqual(job_run.job_type, JOB_TYPE)
        self.assertEqual(job_run.status, JobRun.Status.SUCCESS)
        self.assertEqual(job_run.result_json["summary"]["classes_checked"], 1)
        self.assertEqual(job_run.result_json["summary"]["subjects_checked"], 1)
        self.assertEqual(job_run.result_json["summary"]["with_problems"], 1)
        self.assertEqual(job_run.result_json["report"]["status"], "not_configured")
        self.assertEqual(len(job_run.result_json["rows"]), 1)
        self.assertTrue(job_run.logs.filter(message="Class check started").exists())
        self.assertTrue(job_run.logs.filter(message="Workbook downloaded").exists())
        self.assertTrue(job_run.logs.filter(message="Students found").exists())
        self.assertTrue(job_run.logs.filter(message="Filled criteria found").exists())
        self.assertTrue(job_run.logs.filter(message="Grades checked").exists())
        self.assertTrue(job_run.logs.filter(message="Missing grades found").exists())
        self.assertTrue(job_run.logs.filter(message="Sheet checked").exists())
        self.assertTrue(job_run.logs.filter(message="Problems found").exists())
        self.assertTrue(job_run.logs.filter(message="Descriptor/criteria fill check finished").exists())

    @patch("validation.descriptor_criteria_report._write_payload", side_effect=RuntimeError("write denied"))
    @patch("validation.descriptor_criteria_report._build_sheets_service")
    def test_report_failure_keeps_check_result_and_marks_job_partial(self, _mocked_service, _mocked_write):
        DescriptorCriteriaReportTarget.objects.create(
            name="Report",
            google_sheet_url="https://docs.google.com/spreadsheets/d/report123/edit",
            is_active=True,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "source.xlsx"
            _build_descriptor_criteria_workbook(source_path)

            with patch("validation.descriptor_criteria_fill.fetch_workbook_for_link", return_value=source_path):
                job_run = run_descriptor_criteria_fill_check_job(class_code="7A")

        self.assertEqual(job_run.status, JobRun.Status.PARTIAL)
        self.assertEqual(job_run.result_json["summary"]["subjects_checked"], 1)
        self.assertEqual(job_run.result_json["report"]["status"], "failed")
        self.assertTrue(job_run.logs.filter(message="Report update failed").exists())

    def test_enqueue_creates_pending_job_and_starts_worker(self):
        with patch("validation.descriptor_criteria_fill.threading.Thread") as thread_cls:
            job_run = enqueue_descriptor_criteria_fill_check_job(class_code="7A")

        self.assertEqual(job_run.status, JobRun.Status.PENDING)
        self.assertEqual(job_run.params_json, {"class_code": "7A", "all_active": True, "trigger": "manual"})
        self.assertTrue(job_run.logs.filter(message="Descriptor/criteria fill check queued").exists())
        thread_cls.assert_called_once()
        thread_cls.return_value.start.assert_called_once()


class DescriptorCriteriaFillViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="checker", password="p")
        self.user.user_permissions.add(Permission.objects.get(codename="view_classsheetlink"))
        self.user.user_permissions.add(Permission.objects.get(codename="run_check_missing_data"))
        self.client.force_login(self.user)
        self.link = ClassSheetLink.objects.create(
            class_code="7A",
            subject_name="Math",
            teacher_name="Teacher A",
            google_sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit",
            is_active=True,
        )

    def test_list_links_contains_descriptor_criteria_check_button(self):
        response = self.client.get(reverse("journal_links:list_links"))

        self.assertContains(response, "Проверить дескрипторы, критерии и оценки")

    def test_post_runs_check_for_selected_class(self):
        fake_job = JobRun.objects.create(job_type=JOB_TYPE, result_json={"summary": {}, "rows": []})

        with patch(
            "journal_links.views.enqueue_descriptor_criteria_fill_check_job",
            return_value=fake_job,
        ) as mocked:
            response = self.client.post(
                reverse("journal_links:descriptor_criteria_fill_check"),
                {"class_code": "7A"},
            )

        self.assertEqual(response.status_code, 302)
        self.assertIn(str(fake_job.id), response.url)
        mocked.assert_called_once_with(class_code="7A", all_active=False, initiated_by=self.user, trigger="manual")

    def test_run_now_works_when_schedule_disabled(self):
        schedule = DescriptorCriteriaCheckSchedule.load()
        schedule.is_enabled = False
        schedule.save()
        fake_job = JobRun.objects.create(job_type=JOB_TYPE, result_json={"summary": {}, "rows": []})

        with patch(
            "journal_links.views.enqueue_descriptor_criteria_fill_check_job",
            return_value=fake_job,
        ) as mocked:
            response = self.client.post(
                reverse("journal_links:descriptor_criteria_fill_check"),
                {"action": "run_now"},
            )

        self.assertEqual(response.status_code, 302)
        self.assertIn(str(fake_job.id), response.url)
        mocked.assert_called_once_with(class_code=None, all_active=True, initiated_by=self.user, trigger="manual")

    def test_schedule_toggle_saves_enabled_state(self):
        response = self.client.post(
            reverse("journal_links:descriptor_criteria_fill_check"),
            {
                "action": "save_schedule",
                "is_enabled": "on",
                "interval_minutes": "30",
                "active_job_timeout_minutes": "120",
            },
        )

        self.assertEqual(response.status_code, 302)
        schedule = DescriptorCriteriaCheckSchedule.load()
        self.assertTrue(schedule.is_enabled)
        self.assertEqual(schedule.interval_minutes, 30)
        self.assertEqual(schedule.active_job_timeout_minutes, 120)
        self.assertEqual(schedule.updated_by, self.user)
        self.assertIsNotNone(schedule.next_run_at)

    def test_schedule_interval_saves_and_reschedules_next_run(self):
        schedule = DescriptorCriteriaCheckSchedule.load()
        schedule.is_enabled = True
        schedule.interval_minutes = 90
        schedule.next_run_at = timezone.now() + timedelta(minutes=80)
        schedule.save()

        now = timezone.now()
        with patch("journal_links.views.timezone.now", return_value=now):
            response = self.client.post(
                reverse("journal_links:descriptor_criteria_fill_check"),
                {
                    "action": "save_schedule",
                    "is_enabled": "on",
                    "interval_minutes": "45",
                    "active_job_timeout_minutes": "120",
                },
            )

        self.assertEqual(response.status_code, 302)
        schedule.refresh_from_db()
        self.assertEqual(schedule.interval_minutes, 45)
        self.assertEqual(schedule.next_run_at, now + timedelta(minutes=45))

    def test_schedule_block_renders(self):
        response = self.client.get(reverse("journal_links:descriptor_criteria_fill_check"))

        self.assertContains(response, "Автопроверка дескрипторов, критериев и оценок")
        self.assertContains(response, "Включить автопроверку")
        self.assertContains(response, "Интервал, минут")
        self.assertContains(response, "Таймаут зависшего запуска")
        self.assertContains(response, "Запустить сейчас")

    def test_page_renders_summary_table_and_filters(self):
        job_run = JobRun.objects.create(
            job_type=JOB_TYPE,
            result_json={
                "summary": {
                    "classes_checked": 1,
                    "subjects_checked": 1,
                    "fully_filled": 1,
                    "with_problems": 0,
                },
                "rows": [
                    {
                        "class_code": "7A",
                        "subject_name": "Math",
                        "teacher_name": "Teacher A",
                        "module_number": 1,
                        "descriptor_status": "filled",
                        "criteria_total": 2,
                        "criteria_filled": 2,
                        "criteria_missing": 0,
                        "grades_total": 2,
                        "grades_filled": 2,
                        "grades_missing": 0,
                        "grades_ratio": "2/2",
                        "grades_status": "ok",
                        "overall_status": "ok",
                        "sheet_url": "https://docs.google.com/spreadsheets/d/abc123/edit",
                    }
                ],
            },
        )

        response = self.client.get(
            reverse("journal_links:descriptor_criteria_fill_check"),
            {"run_id": job_run.id, "status": "ok"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Классов проверено")
        self.assertContains(response, "Предметов проверено")
        self.assertContains(response, "Полностью заполнено")
        self.assertContains(response, "Math")
        self.assertContains(response, "Teacher A")
        self.assertContains(response, "2 / 2")
        self.assertContains(response, "Оценки")
        self.assertContains(response, "2/2")
