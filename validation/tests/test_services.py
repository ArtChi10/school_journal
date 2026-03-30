import json
import tempfile
from pathlib import Path

from django.test import SimpleTestCase
from openpyxl import Workbook

from validation.services import WorkbookReadError, parse_subject_sheet, validate_sheet, validate_workbook

ALLOWED_DESCRIPTOR = "Выполняет самостоятельно | Independent"


def _build_valid_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade5A"
    ws.cell(row=1, column=2, value="Класс | Grade")
    ws.cell(row=1, column=3, value="5A")
    ws.cell(row=2, column=2, value="Учитель | Teacher")
    ws.cell(row=2, column=3, value="Ms Doe")
    ws.cell(row=3, column=2, value="Модуль | Module")
    ws.cell(row=3, column=3, value="1")
    ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
    ws.cell(row=4, column=3, value="Term descriptor")
    ws.cell(row=5, column=1, value="Имя")
    ws.cell(row=5, column=2, value="Фамилия")
    ws.cell(row=5, column=3, value="Критерий 1")
    ws.cell(row=5, column=4, value="Критерий 2")
    ws.cell(row=5, column=5, value="Тест 1")
    ws.cell(row=5, column=6, value="Комментарий")
    ws.cell(row=5, column=7, value="Пересдача")

    ws.cell(row=7, column=1, value="Иван")
    ws.cell(row=7, column=2, value="Иванов")
    ws.cell(row=7, column=3, value=ALLOWED_DESCRIPTOR)
    ws.cell(row=7, column=4, value=ALLOWED_DESCRIPTOR)
    ws.cell(row=7, column=5, value=75)

    wb.save(path)


def _build_problem_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade5B"

    ws.cell(row=5, column=1, value="Имя")
    ws.cell(row=5, column=2, value="Фамилия")
    ws.cell(row=5, column=3, value="Критерий 1")
    ws.cell(row=5, column=4, value="Тест 1")
    ws.cell(row=5, column=5, value="Тест 2")
    ws.cell(row=5, column=6, value="Комментарий")
    ws.cell(row=5, column=7, value="Пересдача")

    ws.cell(row=7, column=1, value="Петр")
    ws.cell(row=7, column=2, value="Петров")
    ws.cell(row=7, column=3, value="invalid descriptor")
    ws.cell(row=7, column=4, value="abc")
    ws.cell(row=7, column=5, value=20)

    wb.save(path)


class ValidateWorkbookTests(SimpleTestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._tmpdir = tempfile.TemporaryDirectory()
        tmp_path = Path(cls._tmpdir.name)
        cls.valid_file = tmp_path / "valid_workbook.xlsx"
        cls.problem_file = tmp_path / "problem_workbook.xlsx"

        _build_valid_workbook(cls.valid_file)
        _build_problem_workbook(cls.problem_file)

    @classmethod
    def tearDownClass(cls):
        cls._tmpdir.cleanup()
        super().tearDownClass()

    def test_valid_workbook_returns_unified_json_shape(self):
        result = validate_workbook(str(self.valid_file))

        self.assertEqual(set(result.keys()), {"summary", "issues"})
        self.assertEqual(
            set(result["summary"].keys()),
            {"total", "critical", "warning", "info", "sheets_skipped"},
        )
        self.assertEqual(
            result["summary"],
            {"total": 0, "critical": 0, "warning": 0, "info": 0, "sheets_skipped": 0},
        )
        self.assertEqual(result["issues"], [])

        # ручная проверка JSON-формата (сериализация без ошибок)
        payload = json.dumps(result, ensure_ascii=False)
        self.assertIn('"summary"', payload)
        self.assertIn('"issues"', payload)

    def test_problem_workbook_uses_unified_issue_schema(self):
        result = validate_workbook(str(self.problem_file))

        self.assertGreater(result["summary"]["total"], 0)
        self.assertEqual(result["summary"]["critical"], 5)
        self.assertEqual(result["summary"]["warning"], 3)
        self.assertEqual(result["summary"]["info"], 0)
        self.assertEqual(result["summary"]["sheets_skipped"], 0)

        self.assertEqual(len(result["issues"]), result["summary"]["total"])

        required_issue_keys = {
            "code",
            "severity",
            "sheet",
            "row",
            "student",
            "field",
            "message",
            "class_code",
            "subject_name",
            "teacher_name",
            "module_number",
            "column_type",
        }
        for issue in result["issues"]:
            self.assertEqual(set(issue.keys()), required_issue_keys)

    def test_issue_payload_contains_subject_teacher_context(self):
        result = validate_workbook(str(self.problem_file))
        self.assertGreater(len(result["issues"]), 0)

        issue = result["issues"][0]
        self.assertIn("class_code", issue)
        self.assertIn("subject_name", issue)
        self.assertIn("teacher_name", issue)
        self.assertIn("module_number", issue)
        self.assertIn("column_type", issue)
        self.assertEqual(issue["subject_name"], "Grade5B")

        column_types = {item["column_type"] for item in result["issues"] if item["column_type"] is not None}
        self.assertTrue(column_types.issubset({"criterion", "test", "comment", "retake"}))

    def test_read_errors_raise_predictable_exception(self):
        with self.assertRaises(WorkbookReadError):
            validate_workbook("not_existing_workbook.xlsx")

    def test_invalid_xlsx_content_raises_predictable_exception(self):
        broken_file = Path(self._tmpdir.name) / "broken.xlsx"
        broken_file.write_text("not an xlsx file", encoding="utf-8")

        with self.assertRaises(WorkbookReadError):
            validate_workbook(str(broken_file))

    def test_tutor_sheet_skipped(self):
        workbook_file = Path(self._tmpdir.name) / "tutor_skipped.xlsx"
        wb = Workbook()

        ws_subject = wb.active
        ws_subject.title = "Math"
        ws_subject.cell(row=5, column=1, value="Имя")
        ws_subject.cell(row=5, column=2, value="Фамилия")
        ws_subject.cell(row=5, column=3, value="Критерий 1")
        ws_subject.cell(row=6, column=1, value="Иван")
        ws_subject.cell(row=6, column=2, value="Иванов")
        ws_subject.cell(row=6, column=3, value=ALLOWED_DESCRIPTOR)

        ws_tutor = wb.create_sheet("Тьютор | Tutor")
        ws_tutor.cell(row=5, column=1, value="Имя")
        ws_tutor.cell(row=5, column=2, value="Фамилия")
        ws_tutor.cell(row=5, column=3, value="Критерий 1")
        ws_tutor.cell(row=6, column=1, value="Анна")
        ws_tutor.cell(row=6, column=2, value="Петрова")
        ws_tutor.cell(row=6, column=3, value="")

        wb.save(workbook_file)

        with self.assertLogs("validation.services", level="INFO") as captured_logs:
            result = validate_workbook(str(workbook_file))

        self.assertEqual(result["summary"]["critical"], 0)
        self.assertEqual(result["summary"]["warning"], 0)
        self.assertEqual(result["summary"]["sheets_skipped"], 1)
        self.assertFalse(any(issue["sheet"] == "Тьютор | Tutor" for issue in result["issues"]))
        self.assertTrue(any("Skipping sheet 'Тьютор | Tutor' with type 'tutor'" in log for log in captured_logs.output))

    def test_workbook_with_tutor_and_subject_sheets(self):
        workbook_file = Path(self._tmpdir.name) / "tutor_and_subject.xlsx"
        wb = Workbook()

        ws_subject = wb.active
        ws_subject.title = "Biology"
        ws_subject.cell(row=5, column=1, value="Имя")
        ws_subject.cell(row=5, column=2, value="Фамилия")
        ws_subject.cell(row=5, column=3, value="Критерий 1")
        ws_subject.cell(row=5, column=4, value="Тест 1")
        ws_subject.cell(row=5, column=5, value="Комментарий")
        ws_subject.cell(row=5, column=6, value="Пересдача")
        ws_subject.cell(row=6, column=1, value="Олег")
        ws_subject.cell(row=6, column=2, value="Сергеев")
        ws_subject.cell(row=6, column=3, value=ALLOWED_DESCRIPTOR)
        ws_subject.cell(row=6, column=4, value=80)

        ws_tutor = wb.create_sheet("Tutor support")
        ws_tutor.cell(row=5, column=1, value="Имя")
        ws_tutor.cell(row=5, column=2, value="Фамилия")
        ws_tutor.cell(row=5, column=3, value="Критерий 1")
        ws_tutor.cell(row=6, column=1, value="Анна")
        ws_tutor.cell(row=6, column=2, value="Петрова")
        ws_tutor.cell(row=6, column=3, value="")

        wb.save(workbook_file)

        result = validate_workbook(str(workbook_file))

        self.assertEqual(result["summary"]["total"], 0)
        self.assertEqual(result["summary"]["critical"], 0)
        self.assertEqual(result["summary"]["warning"], 0)
        self.assertEqual(result["summary"]["sheets_skipped"], 1)
class ParseSubjectSheetTests(SimpleTestCase):
    def test_parses_metadata_criteria_students_and_values_dynamically(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Math"

        ws.cell(row=1, column=2, value="Класс | Grade")
        ws.cell(row=1, column=3, value="5A")
        ws.cell(row=2, column=2, value="Учитель | Teacher")
        ws.cell(row=2, column=3, value="Ms Doe")
        ws.cell(row=3, column=2, value="Модуль | Module")
        ws.cell(row=3, column=3, value="3")
        ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
        ws.cell(row=4, column=3, value="Term descriptor")

        ws.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
        ws.cell(row=6, column=3, value="Критерий 1")
        ws.cell(row=6, column=4, value="Критерий 2")
        ws.cell(row=6, column=5, value="Quiz 1")
        ws.cell(row=6, column=6, value="Comment")
        ws.cell(row=6, column=7, value="Retake")
        ws.cell(row=7, column=1, value="ИМЯ")
        ws.cell(row=7, column=2, value="ФАМИЛИЯ")

        ws.cell(row=8, column=1, value="Иван")
        ws.cell(row=8, column=2, value="Иванов")
        ws.cell(row=8, column=3, value=ALLOWED_DESCRIPTOR)
        ws.cell(row=8, column=4, value=ALLOWED_DESCRIPTOR)
        ws.cell(row=8, column=5, value=88)
        ws.cell(row=8, column=6, value="Good")
        ws.cell(row=8, column=7, value="-")

        parsed = parse_subject_sheet(ws)

        self.assertEqual(parsed["metadata"]["class"], "5A")
        self.assertEqual(parsed["metadata"]["teacher"], "Ms Doe")
        self.assertEqual(parsed["metadata"]["module"], "3")
        self.assertEqual(parsed["metadata"]["descriptor"], "Term descriptor")
        self.assertEqual(parsed["criteria_cols"], [3, 4])
        self.assertEqual(parsed["test_cols"], [5])
        self.assertEqual(parsed["comment_col"], 6)
        self.assertEqual(parsed["retake_col"], 7)
        self.assertEqual(parsed["student_header_row"], 7)
        self.assertEqual(parsed["student_start_row"], 8)
        self.assertEqual(len(parsed["students"]), 1)
        self.assertEqual(parsed["students"][0]["student"], "Иван Иванов")
        self.assertEqual(parsed["students"][0]["criteria_values"]["Критерий 1"], ALLOWED_DESCRIPTOR)
        self.assertEqual(parsed["students"][0]["test_values"]["Quiz 1"], 88)

    def test_detect_criteria_row_dynamic(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "History"

        ws.cell(row=9, column=2, value="Критерии оценивания | Assessment criteria")
        ws.cell(row=9, column=3, value="Критерий 1")
        ws.cell(row=9, column=4, value="Quiz 1")
        ws.cell(row=10, column=1, value="ИМЯ")
        ws.cell(row=10, column=2, value="ФАМИЛИЯ")
        ws.cell(row=11, column=1, value="Алия")
        ws.cell(row=11, column=2, value="Садыкова")
        ws.cell(row=11, column=3, value=ALLOWED_DESCRIPTOR)
        ws.cell(row=11, column=4, value=81)

        parsed = parse_subject_sheet(ws)

        self.assertEqual(parsed["criteria_row"], 9)

    def test_detect_student_start_dynamic(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Science"

        ws.cell(row=4, column=2, value="Критерии оценивания | Assessment criteria")
        ws.cell(row=4, column=3, value="Критерий 1")
        ws.cell(row=4, column=4, value="Тест 1")
        ws.cell(row=6, column=1, value="ИМЯ")
        ws.cell(row=6, column=2, value="ФАМИЛИЯ")
        ws.cell(row=7, column=1, value="Петр")
        ws.cell(row=7, column=2, value="Петров")
        ws.cell(row=7, column=3, value=ALLOWED_DESCRIPTOR)
        ws.cell(row=7, column=4, value=81)

        parsed = parse_subject_sheet(ws)

        self.assertEqual(parsed["student_header_row"], 6)
        self.assertEqual(parsed["student_start_row"], 7)
        self.assertEqual(parsed["students"][0]["row"], 7)

    def test_real_data_area_trimmed(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Math"

        ws.cell(row=3, column=2, value="Критерии оценивания | Assessment criteria")
        ws.cell(row=3, column=3, value="Критерий 1")
        ws.cell(row=3, column=4, value="Quiz 1")
        ws.cell(row=4, column=1, value="ИМЯ")
        ws.cell(row=4, column=2, value="ФАМИЛИЯ")
        ws.cell(row=5, column=1, value="Иван")
        ws.cell(row=5, column=2, value="Иванов")
        ws.cell(row=5, column=3, value=ALLOWED_DESCRIPTOR)
        ws.cell(row=5, column=4, value=81)
        # Emulate formatting-only tails: grow worksheet dimensions without real values
        ws.cell(row=200, column=30)

        issues = validate_sheet(ws, "Math")

        self.assertFalse(any(i.code == "EMPTY_CRITERION" for i in issues))
        self.assertFalse(any(i.code == "INVALID_CRITERION_VALUE" for i in issues))

class ValidateSheetColumnTypeTests(SimpleTestCase):
    def _build_mixed_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Mixed"
        ws.cell(row=5, column=1, value="Имя")
        ws.cell(row=5, column=2, value="Фамилия")
        ws.cell(row=5, column=3, value="Критерий 1")
        ws.cell(row=5, column=4, value="Тест 1")
        ws.cell(row=5, column=5, value="Комментарий")
        ws.cell(row=5, column=6, value="Пересдача")
        ws.cell(row=6, column=1, value="Алия")
        ws.cell(row=6, column=2, value="Садыкова")
        return ws

    def test_criterion_column_rejects_numeric(self):
        ws = self._build_mixed_sheet()
        ws.cell(row=6, column=3, value=72)
        ws.cell(row=6, column=4, value=80)

        issues = validate_sheet(ws, "Mixed")

        self.assertTrue(any(i.code == "CRITERION_EXPECTS_LEVEL" for i in issues))
        self.assertFalse(any(i.code == "TEST_SCORE_NOT_NUMERIC" for i in issues))

    def test_test_column_accepts_numeric(self):
        ws = self._build_mixed_sheet()
        ws.cell(row=6, column=3, value=ALLOWED_DESCRIPTOR)
        ws.cell(row=6, column=4, value=72)

        issues = validate_sheet(ws, "Mixed")

        self.assertFalse(any(i.code in {"TEST_SCORE_NOT_NUMERIC", "TEST_SCORE_OUT_OF_RANGE"} for i in issues))

    def test_invalid_retake_value(self):
        ws = self._build_mixed_sheet()
        ws.cell(row=6, column=3, value=ALLOWED_DESCRIPTOR)
        ws.cell(row=6, column=4, value=30)
        ws.cell(row=6, column=5, value="Нужна работа")
        ws.cell(row=6, column=6, value="Maybe")

        issues = validate_sheet(ws, "Mixed")

        self.assertTrue(any(i.code == "INVALID_RETAKE_VALUE" for i in issues))


class ValidateLowScoreRulesTests(SimpleTestCase):
    def _build_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "LowScore"
        ws.cell(row=5, column=1, value="Имя")
        ws.cell(row=5, column=2, value="Фамилия")
        ws.cell(row=5, column=3, value="Критерий 1")
        ws.cell(row=5, column=4, value="Тест 1")
        ws.cell(row=5, column=5, value="Тест 2")
        ws.cell(row=5, column=6, value="Комментарий")
        ws.cell(row=5, column=7, value="Пересдача")
        ws.cell(row=6, column=1, value="Иван")
        ws.cell(row=6, column=2, value="Иванов")
        ws.cell(row=6, column=3, value=ALLOWED_DESCRIPTOR)
        return ws

    def test_comment_required_when_low_score(self):
        ws = self._build_sheet()
        ws.cell(row=6, column=4, value=49)
        ws.cell(row=6, column=5, value=51)
        ws.cell(row=6, column=6, value="")
        ws.cell(row=6, column=7, value="да")

        issues = validate_sheet(ws, "LowScore")

        comment_issues = [i for i in issues if i.code == "COMMENT_REQUIRED"]
        self.assertEqual(len(comment_issues), 1)
        self.assertIn("col_4", comment_issues[0].message)
        self.assertIn("49", comment_issues[0].message)

    def test_retake_required_when_low_score(self):
        ws = self._build_sheet()
        ws.cell(row=6, column=4, value=49)
        ws.cell(row=6, column=6, value="Есть прогресс")
        ws.cell(row=6, column=7, value="")

        issues = validate_sheet(ws, "LowScore")

        retake_issues = [i for i in issues if i.code == "RETAKE_REQUIRED"]
        self.assertEqual(len(retake_issues), 1)
        self.assertIn("col_4", retake_issues[0].message)
        self.assertIn("49", retake_issues[0].message)

    def test_no_low_score_no_comment_retake_requirement(self):
        ws = self._build_sheet()
        ws.cell(row=6, column=4, value=51)
        ws.cell(row=6, column=5, value=80)
        ws.cell(row=6, column=6, value="")
        ws.cell(row=6, column=7, value="")

        issues = validate_sheet(ws, "LowScore")

        self.assertFalse(any(i.code == "COMMENT_REQUIRED" for i in issues))
        self.assertFalse(any(i.code == "RETAKE_REQUIRED" for i in issues))


class ValidateSheetMetadataTests(SimpleTestCase):
    def _build_sheet_with_meta(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Meta"
        ws.cell(row=1, column=2, value="Класс | Grade")
        ws.cell(row=1, column=3, value="5A")
        ws.cell(row=2, column=2, value="Учитель | Teacher")
        ws.cell(row=2, column=3, value="Ms Doe")
        ws.cell(row=3, column=2, value="Модуль | Module")
        ws.cell(row=3, column=3, value="1")
        ws.cell(row=4, column=2, value="Дескриптор | Descriptor")
        ws.cell(row=4, column=3, value="Term descriptor")

        ws.cell(row=6, column=2, value="Критерии оценивания | Assessment criteria")
        ws.cell(row=6, column=3, value="Критерий 1")
        ws.cell(row=7, column=1, value="Имя")
        ws.cell(row=7, column=2, value="Фамилия")
        return ws

    def test_missing_teacher_meta(self):
        ws = self._build_sheet_with_meta()
        ws.cell(row=2, column=3, value="")

        issues = validate_sheet(ws, "Meta")

        teacher_issues = [i for i in issues if i.code == "MISSING_TEACHER_META"]
        self.assertEqual(len(teacher_issues), 1)
        self.assertEqual(teacher_issues[0].severity, "critical")
        self.assertEqual(teacher_issues[0].row, 2)

    def test_invalid_module_meta(self):
        ws = self._build_sheet_with_meta()
        ws.cell(row=3, column=3, value="module-one")

        issues = validate_sheet(ws, "Meta")

        module_issues = [i for i in issues if i.code == "INVALID_MODULE_META"]
        self.assertEqual(len(module_issues), 1)
        self.assertEqual(module_issues[0].severity, "warning")
        self.assertEqual(module_issues[0].row, 3)

    def test_missing_descriptor_meta(self):
        ws = self._build_sheet_with_meta()
        ws.cell(row=4, column=3, value="")

        issues = validate_sheet(ws, "Meta")

        descriptor_issues = [i for i in issues if i.code == "MISSING_DESCRIPTOR_META"]
        self.assertEqual(len(descriptor_issues), 1)
        self.assertEqual(descriptor_issues[0].severity, "critical")
        self.assertEqual(descriptor_issues[0].row, 4)