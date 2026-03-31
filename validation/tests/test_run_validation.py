import os
import tempfile
from pathlib import Path
from unittest.mock import patch

from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from openpyxl import Workbook

from jobs.models import JobRun
from journal_links.models import ClassSheetLink
from validation.job_runner import _build_export_url, fetch_workbook_for_link, run_check_missing_data_job, run_validation_job
from notifications.models import NotificationEvent

ALLOWED_DESCRIPTOR = "Выполняет самостоятельно | Independent"


def _build_valid_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade5A"

    ws.cell(row=5, column=1, value="Имя")
    ws.cell(row=5, column=2, value="Фамилия")
    ws.cell(row=5, column=3, value="Критерий 1")
    ws.cell(row=5, column=4, value="Тест 1")
    ws.cell(row=5, column=5, value="Комментарий")
    ws.cell(row=5, column=6, value="Пересдача")

    ws.cell(row=7, column=1, value="Иван")
    ws.cell(row=7, column=2, value="Иванов")
    ws.cell(row=7, column=3, value=ALLOWED_DESCRIPTOR)
    ws.cell(row=7, column=4, value=95)
    ws.cell(row=7, column=5, value="ok")
    ws.cell(row=7, column=6, value="no")

    wb.save(path)

class FetchWorkbookAccessModeTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.link = ClassSheetLink.objects.create(
            class_code="4A",
            subject_name="Math",
            teacher_name="Teacher",
            google_sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit#gid=0",
            is_active=True,
        )

    def test_fetch_workbook_uses_oauth_owner_mode(self):
        with patch.dict(
            os.environ,
            {
                "GOOGLE_ACCESS_MODE": "oauth_owner",
                "GOOGLE_OAUTH_CLIENT_SECRET_PATH": "/tmp/client_secret.json",
                "GOOGLE_OAUTH_TOKEN_PATH": "/tmp/token.json",
            },
            clear=False,
        ):
            with patch("validation.job_runner._download_workbook_oauth_owner") as oauth_mock:
                oauth_mock.return_value = Path("/tmp/test.xlsx")
                path = fetch_workbook_for_link(self.link)

        self.assertEqual(path, Path("/tmp/test.xlsx"))
        oauth_mock.assert_called_once_with(self.link)

    def test_fetch_workbook_oauth_owner_requires_token_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            client_secret = Path(tmpdir) / "client_secret.json"
            client_secret.write_text("{}", encoding="utf-8")

            with patch.dict(
                os.environ,
                {
                    "GOOGLE_ACCESS_MODE": "oauth_owner",
                    "GOOGLE_OAUTH_CLIENT_SECRET_PATH": str(client_secret),
                    "GOOGLE_OAUTH_TOKEN_PATH": str(Path(tmpdir) / "missing_token.json"),
                },
                clear=False,
            ):
                with self.assertRaisesRegex(Exception, "GOOGLE_OAUTH_TOKEN_PATH"):
                    fetch_workbook_for_link(self.link)

    def test_build_export_url_ignores_gid_and_exports_full_workbook(self):
        export_url = _build_export_url(self.link.google_sheet_url)
        self.assertEqual(export_url, "https://docs.google.com/spreadsheets/d/abc123/export?format=xlsx")
        self.assertNotIn("gid=", export_url)

    def test_build_export_url_keeps_non_google_urls(self):
        url = "https://example.com/file.xlsx"
        self.assertEqual(_build_export_url(url), url)



class RunValidationJobTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.link = ClassSheetLink.objects.create(
            class_code="5A",
            subject_name="Math",
            teacher_name="Teacher",
            google_sheet_url="https://docs.google.com/spreadsheets/d/abc123/edit#gid=0",
            is_active=True,
        )

    def test_run_validation_job_creates_jobrun_logs_and_result_json(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "valid.xlsx"
            _build_valid_workbook(workbook_path)

            with patch("validation.job_runner.fetch_workbook_for_link", return_value=workbook_path):
                job_run = run_validation_job(link_id=self.link.id)

        self.assertEqual(job_run.job_type, "validation")
        self.assertEqual(job_run.status, JobRun.Status.SUCCESS)
        self.assertIsNotNone(job_run.started_at)
        self.assertIsNotNone(job_run.finished_at)

        self.assertIn("summary", job_run.result_json)
        self.assertIn("issues", job_run.result_json)
        self.assertIn("tables", job_run.result_json)
        self.assertEqual(job_run.result_json["summary"]["tables_total"], 1)
        self.assertEqual(job_run.result_json["summary"]["tables_success"], 1)
        self.assertEqual(job_run.result_json["summary"]["sheets_total"], 1)
        self.assertEqual(job_run.result_json["summary"]["sheets_validated"], 1)
        self.assertEqual(job_run.result_json["summary"]["sheets_skipped"], 0)
        self.assertEqual(job_run.result_json["summary"]["students_total"], 1)
        self.assertIsInstance(job_run.result_json["summary"]["issues_by_code"], dict)
        self.assertTrue(job_run.logs.filter(message="sheet_detected").exists())
        self.assertTrue(job_run.logs.filter(message="sheet_validated").exists())
        self.assertGreaterEqual(job_run.logs.count(), 2)

    def test_command_requires_single_selector(self):
        with self.assertRaises(CommandError):
            call_command("run_validation")

        with self.assertRaises(CommandError):
            call_command("run_validation", "--all-active", "--class-code", "5A")

    def test_run_validation_job_keeps_result_when_temp_file_cleanup_fails(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "valid_locked.xlsx"
            _build_valid_workbook(workbook_path)

            with patch("validation.job_runner.fetch_workbook_for_link", return_value=workbook_path):
                with patch("pathlib.Path.unlink", side_effect=PermissionError("file is locked")):
                    job_run = run_validation_job(link_id=self.link.id)

        self.assertEqual(job_run.status, JobRun.Status.SUCCESS)
        self.assertEqual(job_run.result_json["summary"]["tables_total"], 1)
        self.assertEqual(job_run.result_json["summary"]["tables_success"], 1)
        self.assertTrue(job_run.logs.filter(message="Could not remove temporary workbook file").exists())
    @patch("validation.job_runner.send_telegram")
    def test_check_missing_data_job_is_idempotent_by_payload_hash(self, send_telegram_mock):
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "problem_source.xlsx"
            _build_valid_workbook(source_path)

            from openpyxl import load_workbook
            wb = load_workbook(source_path)
            ws = wb.active
            ws.cell(row=4, column=3, value="")  # descriptor empty
            wb.save(source_path)

            generated_paths: list[Path] = []

            def _provide_copy(_link):
                path = Path(tmpdir) / f"problem_{len(generated_paths)}.xlsx"
                path.write_bytes(source_path.read_bytes())
                generated_paths.append(path)
                return path

            with self.settings(ADMIN_LOG_CHAT_ID="999"):
                with patch("validation.job_runner.fetch_workbook_for_link", side_effect=_provide_copy):
                    first = run_check_missing_data_job(all_active=True)
                    second = run_check_missing_data_job(all_active=True)

        self.assertEqual(first.job_type, "check_missing_data")
        self.assertEqual(first.result_json["telegram"]["status"], "sent")
        self.assertEqual(second.result_json["telegram"]["status"], "skipped_duplicate")
        self.assertEqual(send_telegram_mock.call_count, 1)
        self.assertTrue(
            NotificationEvent.objects.filter(
                job_run=first,
                teacher_name="__admin_missing_data__",
                status=NotificationEvent.Status.SENT,
            ).exists()
        )

    @patch("validation.job_runner.send_telegram")
    def test_check_missing_data_job_keeps_result_when_temp_file_cleanup_fails(self, send_telegram_mock):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "valid_missing_cleanup.xlsx"
            _build_valid_workbook(workbook_path)

            with self.settings(ADMIN_LOG_CHAT_ID="999"):
                with patch("validation.job_runner.fetch_workbook_for_link", return_value=workbook_path):
                    with patch("pathlib.Path.unlink", side_effect=PermissionError("file is locked")):
                        job_run = run_check_missing_data_job(all_active=True)

        self.assertEqual(job_run.status, JobRun.Status.SUCCESS)
        self.assertEqual((job_run.result_json or {}).get("telegram", {}).get("status"), "sent")
        self.assertTrue(job_run.logs.filter(message="Could not remove temporary workbook file").exists())