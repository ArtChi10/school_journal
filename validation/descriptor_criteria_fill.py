from __future__ import annotations

import re
import threading
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from django.db import close_old_connections
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from jobs.models import JobLog, JobRun
from jobs.services import log_step
from journal_links.models import ClassSheetLink
from validation.job_runner import fetch_workbook_for_link

JOB_TYPE = "descriptor_criteria_fill_check"


class DescriptorCriteriaFillReadError(ValueError):
    """Raised when a workbook cannot be read for descriptor/criteria checks."""


def _build_job_params(
    *,
    class_code: str | None,
    all_active: bool,
    trigger: str,
    schedule_id: int | None = None,
    interval_minutes: int | None = None,
) -> dict:
    params = {"class_code": class_code, "all_active": all_active, "trigger": trigger}
    if schedule_id is not None:
        params["schedule_id"] = schedule_id
    if interval_minutes is not None:
        params["interval_minutes"] = interval_minutes
    return params


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _is_filled_criteria_value(value: Any) -> bool:
    if _is_empty(value):
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return False
    return re.fullmatch(r"\d+(?:[.,]\d+)?", str(value).strip()) is None


def _normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\xa0", " ")
    text = re.sub(r"[ \t]*\|\s*", " | ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def _classify_sheet_type(sheet_name: str) -> str:
    normalized_name = (sheet_name or "").strip().lower()
    if "тьютор" in normalized_name or "tutor" in normalized_name:
        return "tutor"
    if "служеб" in normalized_name or "service" in normalized_name:
        return "service"
    return "subject"


def _cell_value(ws, row_num: int, col_num: int) -> object:
    cell = ws._cells.get((row_num, col_num))
    return cell.value if cell is not None else None


def _iter_non_empty_cells(ws):
    for (row_num, col_num), cell in sorted(ws._cells.items()):
        if not _is_empty(cell.value):
            yield row_num, col_num, cell.value


def _get_real_data_bounds(ws) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for row_num, col_num, _value in _iter_non_empty_cells(ws):
        max_row = max(max_row, row_num)
        max_col = max(max_col, col_num)
    return max_row, max_col


def _find_anchor(ws, *tokens: str) -> tuple[int, int] | None:
    for row_num, col_num, value in _iter_non_empty_cells(ws):
        normalized = _normalize_text(value)
        if normalized and all(token in normalized for token in tokens):
            return row_num, col_num
    return None


def _cell_right_value(ws, anchor: tuple[int, int] | None) -> object:
    if anchor is None:
        return None
    row_num, col_num = anchor
    return _cell_value(ws, row_num, col_num + 1)


def _parse_module_number(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value

    text = str(value).strip()
    try:
        return int(float(text))
    except (TypeError, ValueError):
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else None


def _find_comment_col(ws, criteria_row: int, start_col: int, max_col: int) -> int | None:
    for col_num in range(start_col, max_col + 1):
        normalized = _normalize_text(_cell_value(ws, criteria_row, col_num))
        if "коммент" in normalized or "comment" in normalized:
            return col_num
    return None


def _find_student_name_header_row(ws, max_row: int) -> int | None:
    for row_num in range(1, max_row + 1):
        normalized = _normalize_text(_cell_value(ws, row_num, 1))
        if normalized in {"имя", "name"}:
            return row_num
    return None


def _student_rows(ws, name_header_row: int | None, max_row: int) -> list[int]:
    if name_header_row is None:
        return []

    rows = []
    for row_num in range(name_header_row + 1, max_row + 1):
        if _is_empty(_cell_value(ws, row_num, 1)):
            break
        rows.append(row_num)
    return rows


def _grade_summary(ws, *, criteria_cols: list[int], student_rows: list[int]) -> dict:
    if not criteria_cols or not student_rows:
        return {
            "grades_total": 0,
            "grades_filled": 0,
            "grades_missing": 0,
            "grades_ratio": "—",
            "grades_status": "not_applicable",
        }

    total = len(criteria_cols) * len(student_rows)
    filled = 0
    for row_num in student_rows:
        for col_num in criteria_cols:
            if not _is_empty(_cell_value(ws, row_num, col_num)):
                filled += 1

    missing = total - filled
    return {
        "grades_total": total,
        "grades_filled": filled,
        "grades_missing": missing,
        "grades_ratio": f"{filled}/{total}",
        "grades_status": "ok" if missing == 0 else "missing",
    }


def check_subject_sheet(ws, *, class_code: str, sheet_url: str) -> dict:
    max_row, max_col = _get_real_data_bounds(ws)
    class_anchor = _find_anchor(ws, "класс", "grade")
    teacher_anchor = _find_anchor(ws, "учитель", "teacher")
    module_anchor = _find_anchor(ws, "module")
    if module_anchor is None:
        module_anchor = _find_anchor(ws, "модуль")
    descriptor_anchor = _find_anchor(ws, "дескриптор", "descriptor")
    criteria_anchor = _find_anchor(ws, "критерии оценивания", "assessment criteria")

    teacher_name = str(_cell_right_value(ws, teacher_anchor) or "").strip()
    module_number = _parse_module_number(_cell_right_value(ws, module_anchor))
    descriptor_value = _cell_right_value(ws, descriptor_anchor)

    if descriptor_anchor is None:
        descriptor_status = "not_found"
    elif _is_empty(descriptor_value):
        descriptor_status = "missing"
    else:
        descriptor_status = "filled"

    criteria_total = 0
    criteria_filled = 0
    criteria_missing = 0
    criteria_status = "not_found"
    criteria_cols_with_text: list[int] = []
    if criteria_anchor is not None:
        criteria_row, criteria_col = criteria_anchor
        comment_col = _find_comment_col(ws, criteria_row, criteria_col + 1, max_col)
        end_col = (comment_col - 1) if comment_col else max_col
        criteria_total = max(0, end_col - criteria_col)
        for col_num in range(criteria_col + 1, end_col + 1):
            value = _cell_value(ws, criteria_row, col_num)
            if not _is_filled_criteria_value(value):
                criteria_missing += 1
            else:
                criteria_filled += 1
                criteria_cols_with_text.append(col_num)
        criteria_status = "filled" if criteria_total > 0 and criteria_missing == 0 else "missing"

    name_header_row = _find_student_name_header_row(ws, max_row)
    student_rows = _student_rows(ws, name_header_row, max_row)
    grade_summary = _grade_summary(ws, criteria_cols=criteria_cols_with_text, student_rows=student_rows)

    overall_status = "ok"
    if descriptor_status != "filled" or criteria_status != "filled" or grade_summary["grades_status"] == "missing":
        overall_status = "problem"

    return {
        "class_code": class_code,
        "sheet_class_code": str(_cell_right_value(ws, class_anchor) or "").strip(),
        "subject_name": ws.title,
        "teacher_name": teacher_name,
        "module_number": module_number,
        "descriptor_status": descriptor_status,
        "criteria_status": criteria_status,
        "criteria_total": criteria_total,
        "criteria_filled": criteria_filled,
        "criteria_missing": criteria_missing,
        "students_total": len(student_rows),
        "grades_total": grade_summary["grades_total"],
        "grades_filled": grade_summary["grades_filled"],
        "grades_missing": grade_summary["grades_missing"],
        "grades_ratio": grade_summary["grades_ratio"],
        "grades_status": grade_summary["grades_status"],
        "overall_status": overall_status,
        "sheet_url": sheet_url,
        "sheet_name": ws.title,
    }


def check_workbook_descriptor_criteria(path: str, *, class_code: str, sheet_url: str) -> dict:
    wb = None
    try:
        wb = load_workbook(path, data_only=True)
    except (FileNotFoundError, InvalidFileException, BadZipFile, OSError) as exc:
        raise DescriptorCriteriaFillReadError(f"Cannot read workbook: {path}") from exc

    rows: list[dict] = []
    sheets_skipped = 0
    sheet_events: list[dict] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_type = _classify_sheet_type(sheet_name)
            sheet_events.append({"event": "sheet_detected", "sheet_name": sheet_name, "sheet_type": sheet_type})
            if sheet_type in {"tutor", "service"}:
                sheets_skipped += 1
                sheet_events.append({"event": "sheet_skipped", "sheet_name": sheet_name, "sheet_type": sheet_type})
                continue

            row = check_subject_sheet(ws, class_code=class_code, sheet_url=sheet_url)
            rows.append(row)
            sheet_events.extend(
                [
                    {
                        "event": "students_found",
                        "sheet_name": sheet_name,
                        "students_total": row["students_total"],
                    },
                    {
                        "event": "filled_criteria_found",
                        "sheet_name": sheet_name,
                        "criteria_filled": row["criteria_filled"],
                    },
                    {
                        "event": "grades_checked",
                        "sheet_name": sheet_name,
                        "grades_total": row["grades_total"],
                        "grades_filled": row["grades_filled"],
                    },
                    {
                        "event": "missing_grades_found",
                        "sheet_name": sheet_name,
                        "grades_missing": row["grades_missing"],
                    },
                ]
            )
            sheet_events.append(
                {
                    "event": "sheet_checked",
                    "sheet_name": sheet_name,
                    "sheet_type": sheet_type,
                    "overall_status": row["overall_status"],
                }
            )

        return {
            "rows": rows,
            "summary": {
                "sheets_total": len(wb.sheetnames),
                "sheets_checked": len(rows),
                "sheets_skipped": sheets_skipped,
                "subjects_ok": sum(1 for row in rows if row["overall_status"] == "ok"),
                "subjects_with_problems": sum(1 for row in rows if row["overall_status"] != "ok"),
            },
            "sheet_events": sheet_events,
        }
    finally:
        if wb is not None:
            wb.close()


def _collect_links(*, class_code: str | None = None, all_active: bool = True) -> list[ClassSheetLink]:
    queryset = ClassSheetLink.objects.filter(is_active=True)
    if class_code:
        queryset = queryset.filter(class_code=class_code)
    elif not all_active:
        queryset = ClassSheetLink.objects.none()
    return list(queryset.order_by("class_code", "id"))


def _update_report_status(job_run: JobRun) -> None:
    from validation.descriptor_criteria_report import update_descriptor_criteria_google_report

    report = update_descriptor_criteria_google_report(job_run)
    result_json = job_run.result_json if isinstance(job_run.result_json, dict) else {}
    result_json["report"] = report
    job_run.result_json = result_json
    update_fields = ["result_json"]
    if report.get("status") == "failed" and job_run.status == JobRun.Status.SUCCESS:
        job_run.status = JobRun.Status.PARTIAL
        update_fields.append("status")
    job_run.save(update_fields=update_fields)


def enqueue_descriptor_criteria_fill_check_job(
    *,
    class_code: str | None = None,
    all_active: bool = True,
    initiated_by=None,
    trigger: str = "manual",
    schedule_id: int | None = None,
    interval_minutes: int | None = None,
) -> JobRun:
    params = _build_job_params(
        class_code=class_code,
        all_active=all_active,
        trigger=trigger,
        schedule_id=schedule_id,
        interval_minutes=interval_minutes,
    )
    job_run = JobRun.objects.create(
        job_type=JOB_TYPE,
        status=JobRun.Status.PENDING,
        started_at=timezone.now(),
        params_json=params,
        initiated_by=initiated_by,
    )
    log_step(
        job_run=job_run,
        level=JobLog.Level.INFO,
        message="Descriptor/criteria fill check queued",
        context=params,
    )
    thread = threading.Thread(
        target=_run_descriptor_criteria_fill_check_thread,
        args=(str(job_run.id), class_code, all_active, trigger, schedule_id, interval_minutes),
        daemon=True,
    )
    thread.start()
    return job_run


def _run_descriptor_criteria_fill_check_thread(
    job_run_id: str,
    class_code: str | None,
    all_active: bool,
    trigger: str,
    schedule_id: int | None,
    interval_minutes: int | None,
) -> None:
    close_old_connections()
    try:
        job_run = JobRun.objects.get(id=job_run_id)
        run_descriptor_criteria_fill_check_job(
            class_code=class_code,
            all_active=all_active,
            job_run=job_run,
            trigger=trigger,
            schedule_id=schedule_id,
            interval_minutes=interval_minutes,
        )
    except Exception as exc:  # noqa: BLE001
        try:
            job_run = JobRun.objects.get(id=job_run_id)
            job_run.status = JobRun.Status.FAILED
            job_run.finished_at = timezone.now()
            job_run.result_json = {"summary": {}, "rows": [], "tables": [], "error": str(exc)}
            job_run.save(update_fields=["status", "finished_at", "result_json"])
            log_step(
                job_run=job_run,
                level=JobLog.Level.ERROR,
                message="Descriptor/criteria fill check background worker failed",
                context={"reason": str(exc)},
            )
        except Exception:
            pass
    finally:
        close_old_connections()


def run_descriptor_criteria_fill_check_job(
    *,
    class_code: str | None = None,
    all_active: bool = True,
    initiated_by=None,
    job_run: JobRun | None = None,
    trigger: str = "manual",
    schedule_id: int | None = None,
    interval_minutes: int | None = None,
) -> JobRun:
    links = _collect_links(class_code=class_code, all_active=all_active)
    params = _build_job_params(
        class_code=class_code,
        all_active=all_active,
        trigger=trigger,
        schedule_id=schedule_id,
        interval_minutes=interval_minutes,
    )
    if job_run is None:
        job_run = JobRun.objects.create(
            job_type=JOB_TYPE,
            status=JobRun.Status.RUNNING,
            started_at=timezone.now(),
            params_json=params,
            initiated_by=initiated_by,
        )
    else:
        job_run.status = JobRun.Status.RUNNING
        job_run.started_at = job_run.started_at or timezone.now()
        job_run.finished_at = None
        job_run.params_json = params
        job_run.result_json = {}
        job_run.save(update_fields=["status", "started_at", "finished_at", "params_json", "result_json"])
    log_step(
        job_run=job_run,
        level=JobLog.Level.INFO,
        message="Descriptor/criteria fill check started",
        context={"links_count": len(links), **params},
    )

    rows: list[dict] = []
    tables: list[dict] = []
    tables_success = 0
    tables_failed = 0
    sheets_total = 0
    sheets_checked = 0
    sheets_skipped = 0
    classes_checked: set[str] = set()

    try:
        for link in links:
            classes_checked.add(link.class_code)
            temp_file: Path | None = None
            log_step(
                job_run=job_run,
                level=JobLog.Level.INFO,
                message="Class check started",
                context={"link_id": link.id, "class_code": link.class_code},
            )
            try:
                temp_file = fetch_workbook_for_link(link)
                log_step(
                    job_run=job_run,
                    level=JobLog.Level.INFO,
                    message="Workbook downloaded",
                    context={"link_id": link.id, "class_code": link.class_code},
                )

                workbook_result = check_workbook_descriptor_criteria(
                    str(temp_file),
                    class_code=link.class_code,
                    sheet_url=link.google_sheet_url,
                )
                table_rows = workbook_result["rows"]
                table_summary = workbook_result["summary"]
                rows.extend(table_rows)
                tables_success += 1
                sheets_total += int(table_summary["sheets_total"])
                sheets_checked += int(table_summary["sheets_checked"])
                sheets_skipped += int(table_summary["sheets_skipped"])

                for event in workbook_result["sheet_events"]:
                    if event["event"] == "sheet_skipped":
                        log_step(
                            job_run=job_run,
                            level=JobLog.Level.INFO,
                            message="Sheet skipped",
                            context={
                                "link_id": link.id,
                                "class_code": link.class_code,
                                "sheet_name": event["sheet_name"],
                                "sheet_type": event["sheet_type"],
                            },
                        )
                    elif event["event"] == "sheet_checked":
                        log_step(
                            job_run=job_run,
                            level=JobLog.Level.INFO,
                            message="Sheet checked",
                            context={
                                "link_id": link.id,
                                "class_code": link.class_code,
                                "sheet_name": event["sheet_name"],
                                "overall_status": event["overall_status"],
                            },
                        )
                    elif event["event"] == "students_found":
                        log_step(
                            job_run=job_run,
                            level=JobLog.Level.INFO,
                            message="Students found",
                            context={
                                "link_id": link.id,
                                "class_code": link.class_code,
                                "sheet_name": event["sheet_name"],
                                "students_total": event["students_total"],
                            },
                        )
                    elif event["event"] == "filled_criteria_found":
                        log_step(
                            job_run=job_run,
                            level=JobLog.Level.INFO,
                            message="Filled criteria found",
                            context={
                                "link_id": link.id,
                                "class_code": link.class_code,
                                "sheet_name": event["sheet_name"],
                                "criteria_filled": event["criteria_filled"],
                            },
                        )
                    elif event["event"] == "grades_checked":
                        log_step(
                            job_run=job_run,
                            level=JobLog.Level.INFO,
                            message="Grades checked",
                            context={
                                "link_id": link.id,
                                "class_code": link.class_code,
                                "sheet_name": event["sheet_name"],
                                "grades_total": event["grades_total"],
                                "grades_filled": event["grades_filled"],
                            },
                        )
                    elif event["event"] == "missing_grades_found":
                        log_step(
                            job_run=job_run,
                            level=JobLog.Level.INFO,
                            message="Missing grades found",
                            context={
                                "link_id": link.id,
                                "class_code": link.class_code,
                                "sheet_name": event["sheet_name"],
                                "grades_missing": event["grades_missing"],
                            },
                        )

                problems_count = sum(1 for row in table_rows if row["overall_status"] != "ok")
                log_step(
                    job_run=job_run,
                    level=JobLog.Level.INFO,
                    message="Problems found",
                    context={"link_id": link.id, "class_code": link.class_code, "problems_count": problems_count},
                )
                tables.append(
                    {
                        "link_id": link.id,
                        "class_code": link.class_code,
                        "status": "success",
                        "summary": table_summary,
                    }
                )
            except Exception as exc:
                tables_failed += 1
                tables.append(
                    {
                        "link_id": link.id,
                        "class_code": link.class_code,
                        "status": "failed",
                        "error": str(exc),
                    }
                )
                log_step(
                    job_run=job_run,
                    level=JobLog.Level.ERROR,
                    message="Class check failed",
                    context={"link_id": link.id, "class_code": link.class_code, "reason": str(exc)},
                )
            finally:
                if temp_file and temp_file.exists():
                    try:
                        temp_file.unlink(missing_ok=True)
                    except PermissionError as exc:
                        log_step(
                            job_run=job_run,
                            level=JobLog.Level.WARNING,
                            message="Could not remove temporary workbook file",
                            context={"path": str(temp_file), "reason": str(exc)},
                        )

        summary = {
            "classes_checked": len(classes_checked),
            "subjects_checked": len(rows),
            "fully_filled": sum(1 for row in rows if row["overall_status"] == "ok"),
            "with_problems": sum(1 for row in rows if row["overall_status"] != "ok"),
            "tables_total": len(links),
            "tables_success": tables_success,
            "tables_failed": tables_failed,
            "sheets_total": sheets_total,
            "sheets_checked": sheets_checked,
            "sheets_skipped": sheets_skipped,
        }
        if not links or tables_success == 0:
            final_status = JobRun.Status.FAILED
        elif tables_failed > 0:
            final_status = JobRun.Status.PARTIAL
        else:
            final_status = JobRun.Status.SUCCESS

        job_run.result_json = {"summary": summary, "rows": rows, "tables": tables}
        job_run.status = final_status
        job_run.finished_at = timezone.now()
        job_run.save(update_fields=["result_json", "status", "finished_at"])
        _update_report_status(job_run)
        log_step(
            job_run=job_run,
            level=JobLog.Level.INFO,
            message="Descriptor/criteria fill check finished",
            context={"status": job_run.status, "summary": summary, "report": job_run.result_json.get("report", {})},
        )
    except Exception as exc:
        job_run.status = JobRun.Status.FAILED
        job_run.finished_at = timezone.now()
        job_run.result_json = {
            "summary": {},
            "rows": rows,
            "tables": tables,
            "error": str(exc),
            "report": {"status": "skipped", "reason": "check_failed"},
        }
        job_run.save(update_fields=["status", "finished_at", "result_json"])
        log_step(
            job_run=job_run,
            level=JobLog.Level.ERROR,
            message="Descriptor/criteria fill check failed",
            context={"reason": str(exc)},
        )

    return job_run
