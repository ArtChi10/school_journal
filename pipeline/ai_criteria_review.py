from __future__ import annotations

import json
import logging
import os
import re
import threading
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from django.db import close_old_connections
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from jobs.models import JobLog, JobRun
from jobs.services import log_step
from journal_links.models import ClassSheetLink
from pipeline.services import _get_openai_client
from validation.job_runner import fetch_workbook_for_link

logger = logging.getLogger(__name__)

JOB_TYPE = "ai_criteria_class_review"
DEFAULT_AI_CRITERIA_REVIEW_MODEL = "gpt-4.1-mini"

AI_CRITERIA_REVIEW_PROMPT = (
    "Ты проверяешь качество формулировок критериев оценивания для школьных предметных листов.\n"
    "Оценивай только критерии, не оценивай учеников, не придумывай оценки и не готовь уведомления.\n"
    "Для каждого критерия определи, можно ли по нему понятно и измеримо оценивать учебный результат.\n"
    "Хороший критерий описывает наблюдаемое действие или результат ученика, без размытых слов вроде "
    "'знает', 'понимает', 'старается', 'хорошо умеет', если они не уточнены измеримым результатом.\n\n"
    "Верни строго JSON без markdown и лишнего текста:\n"
    "{\n"
    '  "criteria": [\n'
    '    {"index": 1, "verdict": "ok|problem", "reason": "короткая причина", "suggested_rewrite": "как улучшить или пустая строка"}\n'
    "  ]\n"
    "}\n\n"
    "Требования:\n"
    "- Сохрани index из входных данных.\n"
    "- Для каждого входного критерия верни ровно один объект.\n"
    "- verdict только ok или problem.\n"
    "- reason обязателен.\n"
    "- suggested_rewrite обязателен; если критерий хороший, верни пустую строку или короткий улучшенный вариант.\n"
)


class AICriteriaReviewError(RuntimeError):
    """Raised when class criteria AI review fails."""


class AICriteriaReviewFormatError(AICriteriaReviewError):
    """Raised when AI returns an invalid batch review payload."""


class AICriteriaWorkbookReadError(ValueError):
    """Raised when a workbook cannot be read for AI criteria review."""


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\xa0", " ")
    text = re.sub(r"[ \t]*\|\s*", " | ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def _is_numeric_only(value: Any) -> bool:
    if _is_empty(value):
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    return re.fullmatch(r"\d+(?:[.,]\d+)?", str(value).strip()) is not None


def _classify_sheet_type(sheet_name: str) -> str:
    normalized = (sheet_name or "").strip().lower()
    if "тьютор" in normalized or "tutor" in normalized:
        return "tutor"
    if "служеб" in normalized or "service" in normalized:
        return "service"
    return "subject"


def _cell_value(ws, row_num: int, col_num: int) -> object:
    cell = ws._cells.get((row_num, col_num))
    return cell.value if cell is not None else None


def _iter_non_empty_cells(ws):
    for (row_num, col_num), cell in sorted(ws._cells.items()):
        if not _is_empty(cell.value):
            yield row_num, col_num, cell.value


def _get_real_data_bounds(ws) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for row_num, col_num, _value in _iter_non_empty_cells(ws):
        max_row = max(max_row, row_num)
        max_col = max(max_col, col_num)
    return max_row, max_col


def _find_anchor(ws, *tokens: str) -> tuple[int, int] | None:
    for row_num, col_num, value in _iter_non_empty_cells(ws):
        normalized = _normalize_text(value)
        if normalized and all(token in normalized for token in tokens):
            return row_num, col_num
    return None


def _cell_right_value(ws, anchor: tuple[int, int] | None) -> object:
    if anchor is None:
        return None
    row_num, col_num = anchor
    return _cell_value(ws, row_num, col_num + 1)


def _parse_module_number(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    try:
        return int(float(text))
    except (TypeError, ValueError):
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else None


def _find_comment_col(ws, criteria_row: int, start_col: int, max_col: int) -> int | None:
    for col_num in range(start_col, max_col + 1):
        normalized = _normalize_text(_cell_value(ws, criteria_row, col_num))
        if "коммент" in normalized or "comment" in normalized:
            return col_num
    return None


def collect_ai_reviewable_criteria_from_workbook(path: str, *, class_code: str, sheet_url: str) -> dict[str, Any]:
    try:
        wb = load_workbook(path, data_only=True)
    except (FileNotFoundError, InvalidFileException, BadZipFile, OSError) as exc:
        logger.error("Failed to read workbook for AI criteria review: %s (%s)", path, exc.__class__.__name__)
        raise AICriteriaWorkbookReadError(f"Cannot read workbook: {path}") from exc

    rows: list[dict[str, Any]] = []
    skipped_empty = 0
    skipped_numeric = 0
    sheets_skipped = 0
    sheets_checked = 0
    seen_keys: set[tuple[str, str, int | None, str]] = set()

    try:
        for sheet_name in wb.sheetnames:
            sheet_type = _classify_sheet_type(sheet_name)
            if sheet_type in {"tutor", "service"}:
                sheets_skipped += 1
                continue

            ws = wb[sheet_name]
            _max_row, max_col = _get_real_data_bounds(ws)
            criteria_anchor = _find_anchor(ws, "критерии оценивания", "assessment criteria")
            if criteria_anchor is None:
                continue

            sheets_checked += 1
            teacher_anchor = _find_anchor(ws, "учитель", "teacher")
            module_anchor = _find_anchor(ws, "module") or _find_anchor(ws, "модуль")
            teacher_name = str(_cell_right_value(ws, teacher_anchor) or "").strip()
            module_number = _parse_module_number(_cell_right_value(ws, module_anchor))
            criteria_row, criteria_col = criteria_anchor
            comment_col = _find_comment_col(ws, criteria_row, criteria_col + 1, max_col)
            end_col = (comment_col - 1) if comment_col else max_col

            for col_num in range(criteria_col + 1, end_col + 1):
                criterion_text = str(_cell_value(ws, criteria_row, col_num) or "").strip()
                if not criterion_text:
                    skipped_empty += 1
                    continue
                if _is_numeric_only(criterion_text):
                    skipped_numeric += 1
                    continue

                dedupe_key = (sheet_name, teacher_name, module_number, _normalize_text(criterion_text))
                if dedupe_key in seen_keys:
                    continue
                seen_keys.add(dedupe_key)
                rows.append(
                    {
                        "class_code": class_code,
                        "subject_name": sheet_name,
                        "teacher_name": teacher_name,
                        "module_number": module_number,
                        "criterion_text": criterion_text,
                        "source_sheet_name": sheet_name,
                        "sheet_url": sheet_url,
                    }
                )
    finally:
        wb.close()

    return {
        "rows": rows,
        "summary": {
            "criteria_collected": len(rows),
            "criteria_skipped_empty": skipped_empty,
            "criteria_skipped_numeric": skipped_numeric,
            "sheets_checked": sheets_checked,
            "sheets_skipped": sheets_skipped,
            "sheets_total": len(wb.sheetnames),
        },
    }


def _request_ai_class_criteria_review(
    criteria: list[dict[str, Any]],
    *,
    ai_client,
    model_name: str,
) -> str:
    payload = {
        "class_code": criteria[0]["class_code"] if criteria else "",
        "criteria": [
            {
                "index": index,
                "subject": row["subject_name"],
                "teacher": row["teacher_name"],
                "module": row["module_number"],
                "criterion": row["criterion_text"],
            }
            for index, row in enumerate(criteria, start=1)
        ],
    }
    response = ai_client.responses.create(
        model=model_name,
        input=[
            {"role": "system", "content": AI_CRITERIA_REVIEW_PROMPT},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ],
        temperature=0,
    )
    return str(getattr(response, "output_text", "") or "").strip()


def parse_ai_class_criteria_response(payload: str, *, expected_count: int) -> dict[int, dict[str, str]]:
    raw = str(payload or "").strip()
    if not raw:
        raise AICriteriaReviewFormatError("AI criteria review returned an empty response")
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise AICriteriaReviewFormatError("AI criteria review returned non-JSON response") from exc

    if isinstance(parsed, dict):
        items = parsed.get("criteria")
    else:
        items = parsed
    if not isinstance(items, list):
        raise AICriteriaReviewFormatError("AI criteria review JSON must contain a criteria list")

    results: dict[int, dict[str, str]] = {}
    for item in items:
        if not isinstance(item, dict):
            raise AICriteriaReviewFormatError("Each AI criteria review item must be an object")
        try:
            index = int(item.get("index"))
        except (TypeError, ValueError) as exc:
            raise AICriteriaReviewFormatError("Each AI criteria review item must include an integer index") from exc
        if index < 1 or index > expected_count:
            continue
        verdict = str(item.get("verdict", "")).strip().lower()
        if verdict in {"valid", "good"}:
            verdict = "ok"
        if verdict in {"invalid", "partial", "bad"}:
            verdict = "problem"
        if verdict not in {"ok", "problem"}:
            raise AICriteriaReviewFormatError("AI verdict must be ok or problem")
        reason = str(item.get("reason", "")).strip()
        suggested_rewrite = str(item.get("suggested_rewrite", item.get("fix", "")) or "").strip()
        if not reason:
            raise AICriteriaReviewFormatError("AI reason is required")
        results[index] = {
            "ai_verdict": verdict,
            "ai_reason": reason,
            "ai_suggested_rewrite": suggested_rewrite,
        }
    return results


def evaluate_class_criteria_with_ai(
    criteria: list[dict[str, Any]],
    *,
    client=None,
    model: str | None = None,
) -> dict[int, dict[str, str]]:
    if not criteria:
        return {}
    ai_client = client or _get_openai_client()
    model_name = model or os.getenv("OPENAI_CRITERIA_CLASS_REVIEW_MODEL") or os.getenv("OPENAI_CRITERIA_MODEL") or DEFAULT_AI_CRITERIA_REVIEW_MODEL
    raw_response = _request_ai_class_criteria_review(criteria, ai_client=ai_client, model_name=model_name)
    return parse_ai_class_criteria_response(raw_response, expected_count=len(criteria))


def _collect_links(*, class_code: str | None, all_active: bool) -> list[ClassSheetLink]:
    queryset = ClassSheetLink.objects.filter(is_active=True)
    if class_code:
        queryset = queryset.filter(class_code=class_code)
    elif not all_active:
        queryset = ClassSheetLink.objects.none()
    return list(queryset.order_by("class_code", "id"))


def _params(*, class_code: str | None, all_active: bool) -> dict[str, Any]:
    return {"class_code": class_code, "all_active": all_active}


def _update_ai_report_status(job_run: JobRun) -> None:
    from pipeline.ai_criteria_report import update_ai_criteria_google_report

    report = update_ai_criteria_google_report(job_run)
    result_json = job_run.result_json if isinstance(job_run.result_json, dict) else {}
    result_json["report"] = report
    job_run.result_json = result_json
    update_fields = ["result_json"]
    if report.get("status") == "failed" and job_run.status == JobRun.Status.SUCCESS:
        job_run.status = JobRun.Status.PARTIAL
        update_fields.append("status")
    job_run.save(update_fields=update_fields)


def enqueue_ai_criteria_class_review_job(
    *,
    class_code: str | None = None,
    all_active: bool = True,
    initiated_by=None,
) -> JobRun:
    params = _params(class_code=class_code, all_active=all_active)
    job_run = JobRun.objects.create(
        job_type=JOB_TYPE,
        status=JobRun.Status.PENDING,
        started_at=timezone.now(),
        params_json=params,
        initiated_by=initiated_by,
    )
    log_step(job_run=job_run, level=JobLog.Level.INFO, message="AI criteria review queued", context=params)
    thread = threading.Thread(
        target=_run_ai_criteria_class_review_thread,
        args=(str(job_run.id), class_code, all_active),
        daemon=True,
    )
    thread.start()
    return job_run


def _run_ai_criteria_class_review_thread(job_run_id: str, class_code: str | None, all_active: bool) -> None:
    close_old_connections()
    try:
        job_run = JobRun.objects.get(id=job_run_id)
        run_ai_criteria_class_review_job(class_code=class_code, all_active=all_active, job_run=job_run)
    except Exception as exc:  # noqa: BLE001
        try:
            job_run = JobRun.objects.get(id=job_run_id)
            job_run.status = JobRun.Status.FAILED
            job_run.finished_at = timezone.now()
            job_run.result_json = {"summary": {}, "rows": [], "tables": [], "error": str(exc)}
            job_run.save(update_fields=["status", "finished_at", "result_json"])
            log_step(job_run=job_run, level=JobLog.Level.ERROR, message="AI criteria review background worker failed", context={"reason": str(exc)})
        except Exception:
            pass
    finally:
        close_old_connections()


def run_ai_criteria_class_review_job(
    *,
    class_code: str | None = None,
    all_active: bool = True,
    initiated_by=None,
    job_run: JobRun | None = None,
    client=None,
    model: str | None = None,
) -> JobRun:
    links = _collect_links(class_code=class_code, all_active=all_active)
    params = _params(class_code=class_code, all_active=all_active)
    if job_run is None:
        job_run = JobRun.objects.create(
            job_type=JOB_TYPE,
            status=JobRun.Status.RUNNING,
            started_at=timezone.now(),
            params_json=params,
            initiated_by=initiated_by,
        )
    else:
        job_run.status = JobRun.Status.RUNNING
        job_run.started_at = job_run.started_at or timezone.now()
        job_run.finished_at = None
        job_run.params_json = params
        job_run.result_json = {}
        job_run.save(update_fields=["status", "started_at", "finished_at", "params_json", "result_json"])

    log_step(job_run=job_run, level=JobLog.Level.INFO, message="AI criteria review started", context={"links_count": len(links), **params})

    rows: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    classes_checked: set[str] = set()
    criteria_sent_to_ai = 0
    criteria_skipped_empty = 0
    criteria_skipped_numeric = 0
    criteria_ok = 0
    criteria_problem = 0
    ai_requests_total = 0
    ai_requests_failed = 0
    tables_success = 0
    tables_failed = 0

    try:
        for link in links:
            classes_checked.add(link.class_code)
            temp_file: Path | None = None
            try:
                log_step(job_run=job_run, level=JobLog.Level.INFO, message="Class criteria collection started", context={"link_id": link.id, "class_code": link.class_code})
                temp_file = fetch_workbook_for_link(link)
                collection = collect_ai_reviewable_criteria_from_workbook(
                    str(temp_file),
                    class_code=link.class_code,
                    sheet_url=link.google_sheet_url,
                )
                criteria = collection["rows"]
                collection_summary = collection["summary"]
                criteria_skipped_empty += int(collection_summary["criteria_skipped_empty"])
                criteria_skipped_numeric += int(collection_summary["criteria_skipped_numeric"])
                tables_success += 1
                log_step(
                    job_run=job_run,
                    level=JobLog.Level.INFO,
                    message="Class criteria collected",
                    context={"link_id": link.id, "class_code": link.class_code, **collection_summary},
                )
                if collection_summary["criteria_skipped_empty"] or collection_summary["criteria_skipped_numeric"]:
                    log_step(
                        job_run=job_run,
                        level=JobLog.Level.INFO,
                        message="Criteria skipped",
                        context={
                            "link_id": link.id,
                            "class_code": link.class_code,
                            "empty": collection_summary["criteria_skipped_empty"],
                            "numeric": collection_summary["criteria_skipped_numeric"],
                        },
                    )

                reviewed_by_index: dict[int, dict[str, str]] = {}
                if criteria:
                    ai_requests_total += 1
                    criteria_sent_to_ai += len(criteria)
                    log_step(job_run=job_run, level=JobLog.Level.INFO, message="AI batch request started", context={"class_code": link.class_code, "criteria_count": len(criteria)})
                    try:
                        reviewed_by_index = evaluate_class_criteria_with_ai(criteria, client=client, model=model)
                        log_step(job_run=job_run, level=JobLog.Level.INFO, message="AI batch request finished", context={"class_code": link.class_code, "results_count": len(reviewed_by_index)})
                    except Exception as exc:  # noqa: BLE001
                        ai_requests_failed += 1
                        log_step(job_run=job_run, level=JobLog.Level.ERROR, message="AI batch request failed", context={"class_code": link.class_code, "reason": str(exc)})

                for index, row in enumerate(criteria, start=1):
                    ai_result = reviewed_by_index.get(index)
                    if ai_result is None:
                        ai_result = {
                            "ai_verdict": "problem",
                            "ai_reason": "AI не вернул результат для этого критерия.",
                            "ai_suggested_rewrite": "",
                        }
                    if ai_result["ai_verdict"] == "ok":
                        criteria_ok += 1
                    else:
                        criteria_problem += 1
                    result_row = {**row, **ai_result}
                    rows.append(result_row)

                tables.append(
                    {
                        "link_id": link.id,
                        "class_code": link.class_code,
                        "status": "success",
                        "criteria_count": len(criteria),
                        "ai_results_count": len(reviewed_by_index),
                        "summary": collection_summary,
                    }
                )
            except Exception as exc:  # noqa: BLE001
                tables_failed += 1
                tables.append({"link_id": link.id, "class_code": link.class_code, "status": "failed", "error": str(exc)})
                log_step(job_run=job_run, level=JobLog.Level.ERROR, message="Class AI criteria review failed", context={"link_id": link.id, "class_code": link.class_code, "reason": str(exc)})
            finally:
                if temp_file and temp_file.exists():
                    try:
                        temp_file.unlink(missing_ok=True)
                    except PermissionError as exc:
                        log_step(job_run=job_run, level=JobLog.Level.WARNING, message="Could not remove temporary workbook file", context={"path": str(temp_file), "reason": str(exc)})

        summary = {
            "classes_checked": len(classes_checked),
            "criteria_sent_to_ai": criteria_sent_to_ai,
            "criteria_skipped_empty": criteria_skipped_empty,
            "criteria_skipped_numeric": criteria_skipped_numeric,
            "criteria_ok": criteria_ok,
            "criteria_problem": criteria_problem,
            "ai_requests_total": ai_requests_total,
            "ai_requests_failed": ai_requests_failed,
            "tables_total": len(links),
            "tables_success": tables_success,
            "tables_failed": tables_failed,
        }
        if not links or (tables_success == 0 and tables_failed > 0):
            final_status = JobRun.Status.FAILED
        elif tables_failed > 0 or ai_requests_failed > 0:
            final_status = JobRun.Status.PARTIAL
        else:
            final_status = JobRun.Status.SUCCESS

        job_run.result_json = {"summary": summary, "rows": rows, "tables": tables}
        job_run.status = final_status
        job_run.finished_at = timezone.now()
        job_run.save(update_fields=["result_json", "status", "finished_at"])
        _update_ai_report_status(job_run)
        log_step(job_run=job_run, level=JobLog.Level.INFO, message="AI criteria review finished", context={"status": job_run.status, "summary": summary, "report": job_run.result_json.get("report", {})})
    except Exception as exc:  # noqa: BLE001
        job_run.status = JobRun.Status.FAILED
        job_run.finished_at = timezone.now()
        job_run.result_json = {
            "summary": {},
            "rows": rows,
            "tables": tables,
            "error": str(exc),
            "report": {"status": "skipped", "reason": "check_failed"},
        }
        job_run.save(update_fields=["status", "finished_at", "result_json"])
        log_step(job_run=job_run, level=JobLog.Level.ERROR, message="AI criteria review failed", context={"reason": str(exc)})

    return job_run
