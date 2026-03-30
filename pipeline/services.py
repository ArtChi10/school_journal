import logging
import os
import re
import json
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)

_CRITERIA_ANCHOR_RU = "критерии оценивания"
_CRITERIA_ANCHOR_EN = "assessment criteria"
_EXCLUDED_HEADER_KEYWORDS = (
    "коммент",
    "comment",
    "пересда",
    "retake",
    "resit",
    "make-up",
)
_DEFAULT_NORMALIZATION_MODEL = "gpt-4.1-mini"
_CRITERION_EVALUATION_PROMPT = (
    "Проанализируй предложенный критерий оценивания для школьного отчета об успеваемости. "
    "Твоя задача — определить, подходит ли он для критериального оценивания по правилам школы.\n\n"
    "Проверяй критерий по этим пунктам:\n"
    "1) Сформулирован ли он в форме глагола 3-го лица и отвечает ли на вопрос «что делает?».\n"
    "2) Является ли глагол конкретным, наблюдаемым и измеримым.\n"
    "3) Описывает ли критерий учебный результат, а не личные качества ребенка, "
    "не поведение вообще и не врожденные способности.\n"
    "4) Соответствует ли критерий содержанию модуля и программе.\n"
    "5) Можно ли по нему реально выставить уровень: «не проявляется», «начальный уровень», "
    "«соответствует ожиданиям», «превосходит ожидания».\n"
    "6) Нет ли в формулировке слишком общих, размытых или субъективных слов вроде: "
    "«знает», «понимает», «умеет хорошо», «старается», «молодец», «способен».\n"
    "7) Нет ли лишней оценочности, эмоциональности или неформального стиля.\n"
    "8) Не относится ли критерий к второстепенным вещам, если он поставлен как один из главных.\n\n"
    "После проверки верни строго JSON-объект БЕЗ markdown и лишнего текста:\n"
    "{\n"
    '  "verdict": "valid|partial|invalid",\n'
    '  "why": "короткое объяснение",\n'
    '  "fix": "что исправить",\n'
    '  "variants": ["вариант 1", "вариант 2"]\n'
    "}\n\n"
    "Требования к формату:\n"
    "- JSON строго валиден.\n"
    "- Поля verdict/why/fix/variants обязательны.\n"
    "- variants — список из 1-3 непустых строк.\n"
    "- verdict только из: valid, partial, invalid.\n\n"
    "Если критерий уже хороший, так и скажи и коротко объясни, почему он удачен."
)

class WorkbookReadError(ValueError):
    """Raised when a workbook cannot be read for criteria extraction."""

class CriterionNormalizationError(ValueError):
    """Raised when criterion normalization via AI fails."""

class CriterionAIFormatError(CriterionNormalizationError):
    """Raised when AI response cannot be parsed into expected structured format."""


def _normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\xa0", " ")
    text = re.sub(r"[ \t]*\|\s*", " | ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def _is_tutor_sheet(sheet_name: str) -> bool:
    title = sheet_name.strip().lower()
    return "тьютор" in title or "tutor" in title


def _find_criteria_header_row(ws) -> int | None:
    for row_num in range(1, ws.max_row + 1):
        anchor = _normalize_text(ws.cell(row=row_num, column=2).value)
        if _CRITERIA_ANCHOR_RU in anchor and _CRITERIA_ANCHOR_EN in anchor:
            return row_num
    return None


def _is_excluded_header(header: object) -> bool:
    normalized = _normalize_text(header)
    return any(keyword in normalized for keyword in _EXCLUDED_HEADER_KEYWORDS)


def _parse_module_number(value: object) -> int:
    if value is None:
        return 0

    if isinstance(value, int):
        return value

    as_text = str(value).strip()
    try:
        return int(float(as_text))
    except (ValueError, TypeError):
        match = re.search(r"\d+", as_text)
        return int(match.group(0)) if match else 0

def _get_openai_client():
    try:
        from openai import OpenAI
    except Exception as exc:  # pragma: no cover - exercised in environments without dependency
        raise CriterionNormalizationError(
            "OpenAI client is not available. Install the 'openai' package."
        ) from exc

    return OpenAI()


def _parse_ai_json_payload(payload: str) -> dict:
    raw = str(payload or "").strip()
    if not raw:
        raise CriterionAIFormatError("AI normalization returned an empty response")

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise CriterionAIFormatError("AI returned non-JSON response") from exc

    if not isinstance(parsed, dict):
        raise CriterionAIFormatError("AI JSON response must be an object")

    verdict = str(parsed.get("verdict", "")).strip().lower()
    why = str(parsed.get("why", "")).strip()
    fix = str(parsed.get("fix", "")).strip()
    variants = parsed.get("variants")

    if verdict not in {"valid", "partial", "invalid"}:
        raise CriterionAIFormatError("AI verdict must be one of: valid, partial, invalid")
    if not why:
        raise CriterionAIFormatError("AI why is required")
    if not fix:
        raise CriterionAIFormatError("AI fix is required")
    if not isinstance(variants, list):
        raise CriterionAIFormatError("AI variants must be a list")

    cleaned_variants = [str(item).strip() for item in variants if str(item).strip()]
    if not cleaned_variants:
        raise CriterionAIFormatError("AI variants must contain at least one non-empty item")

    return {
        "verdict": verdict,
        "why": why,
        "fix": fix,
        "variants": cleaned_variants[:3],
    }


def _request_ai_criterion_evaluation(
    source_text: str,
    *,
    ai_client,
    model_name: str,
) -> str:
    response = ai_client.responses.create(
        model=model_name,
        input=[
            {"role": "system", "content": _CRITERION_EVALUATION_PROMPT},
            {"role": "user", "content": source_text},
        ],
        temperature=0,
    )
    return str(getattr(response, "output_text", "") or "").strip()


def evaluate_criterion_text_with_ai(
    criterion_text: str,
    *,
    client=None,
    model: str | None = None,
) -> dict:
    source_text = str(criterion_text or "").strip()
    if not source_text:
        return {"verdict": "invalid", "why": "", "fix": "", "variants": []}

    ai_client = client or _get_openai_client()
    model_name = model or os.getenv("OPENAI_CRITERIA_MODEL", _DEFAULT_NORMALIZATION_MODEL)

    try:
        raw_response = _request_ai_criterion_evaluation(
            source_text,
            ai_client=ai_client,
            model_name=model_name,
        )
        return _parse_ai_json_payload(raw_response)
    except CriterionAIFormatError:
        try:
            raw_response = _request_ai_criterion_evaluation(
                f"{source_text}\n\nВажно: верни только валидный JSON в заданном формате.",
                ai_client=ai_client,
                model_name=model_name,
            )
            return _parse_ai_json_payload(raw_response)
        except CriterionAIFormatError as exc:
            raise CriterionNormalizationError("AI returned invalid structured format") from exc
    except Exception as exc:
        raise CriterionNormalizationError(
            f"AI normalization failed for criterion '{source_text[:100]}'"
        ) from exc


def normalize_criterion_text_with_ai(
    criterion_text: str,
    *,
    client=None,
    model: str | None = None,
) -> str:
    """Normalize criterion wording via ChatGPT Responses API while preserving intent."""
    source_text = str(criterion_text or "").strip()
    if not source_text:
        return ""

    result = evaluate_criterion_text_with_ai(
        source_text,
        client=client,
        model=model,
    )
    variants = result.get("variants") or []
    return str(variants[0]).strip() if variants else ""

def add_ai_normalized_criteria(
    extracted_rows: list[dict],
    *,
    client=None,
    model: str | None = None,
) -> list[dict]:
    """Return extracted rows enriched with `criterion_text_ai` for each row."""
    normalized_rows: list[dict] = []

    for row in extracted_rows:
        enriched = dict(row)
        enriched["criterion_text_ai"] = normalize_criterion_text_with_ai(
            enriched.get("criterion_text", ""),
            client=client,
            model=model,
        )
        normalized_rows.append(enriched)

    return normalized_rows


def extract_raw_criteria_from_workbook(path: str, class_code: str) -> list[dict]:
    try:
        wb = load_workbook(path, data_only=True)
    except (FileNotFoundError, InvalidFileException, BadZipFile, OSError) as exc:
        logger.error(
            "Failed to read workbook for criteria extraction: %s (%s)",
            path,
            exc.__class__.__name__,
        )
        raise WorkbookReadError(f"Cannot read workbook: {path}") from exc

    records: list[dict] = []
    source_workbook = str(Path(path))

    for sheet_name in wb.sheetnames:
        if _is_tutor_sheet(sheet_name):
            continue

        ws = wb[sheet_name]
        criteria_header_row = _find_criteria_header_row(ws)
        if criteria_header_row is None:
            continue

        subject_name = sheet_name
        teacher_name = str(ws["C2"].value or "").strip()
        module_number = _parse_module_number(ws["C3"].value)

        for col in range(3, ws.max_column + 1):
            header = ws.cell(row=criteria_header_row, column=col).value
            criterion_text = str(header or "").strip()

            if not criterion_text:
                continue
            if _is_excluded_header(header):
                continue

            records.append(
                {
                    "class_code": class_code,
                    "subject_name": subject_name,
                    "teacher_name": teacher_name,
                    "module_number": module_number,
                    "criterion_text": criterion_text,
                    "source_sheet_name": sheet_name,
                    "source_workbook": source_workbook,
                }
            )

    return records