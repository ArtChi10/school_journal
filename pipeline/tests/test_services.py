import json
import tempfile
from pathlib import Path

from django.test import TestCase
from openpyxl import Workbook

from pipeline.models import ValidCriterionTemplate
from pipeline.services import (
    CriterionNormalizationError,
    WorkbookReadError,
    _CRITERION_EVALUATION_PROMPT,
    add_ai_normalized_criteria,
    evaluate_criterion_text_with_ai,
    extract_raw_criteria_from_workbook,
)


def _build_workbook(path: Path) -> None:
    wb = Workbook()

    ws_subject = wb.active
    ws_subject.title = "Math"
    ws_subject["C2"] = "Ms. Frizzle"
    ws_subject["C3"] = "Module 2"

    ws_subject.cell(row=5, column=2, value="Критерии оценивания | \nAssessment criteria")
    ws_subject.cell(row=5, column=3, value="Критерий 1")
    ws_subject.cell(row=5, column=4, value="Критерий 2")
    ws_subject.cell(row=5, column=5, value="Комментарий")
    ws_subject.cell(row=5, column=6, value="Пересдача")
    ws_subject.cell(row=5, column=7, value="Критерий 3")

    ws_tutor = wb.create_sheet("Тьютор | Tutor")
    ws_tutor["C2"] = "Tutor Name"
    ws_tutor["C3"] = 1
    ws_tutor.cell(row=5, column=2, value="Критерии оценивания | \nAssessment criteria")
    ws_tutor.cell(row=5, column=3, value="Не должен попасть")

    ws_no_anchor = wb.create_sheet("History")
    ws_no_anchor["C2"] = "Mr. History"
    ws_no_anchor["C3"] = 3
    ws_no_anchor.cell(row=5, column=3, value="Критерий без якоря")

    wb.save(path)

class _FakeResponses:
    def __init__(self, output_text='{"verdict":"valid","why":"ok","fix":"none","variants":["Нормализованный критерий"]}'):
        self.calls = []
        self.output_text = output_text

    def create(self, **kwargs):
        self.calls.append(kwargs)

        class _Response:
            output_text = ""

        _Response.output_text = self.output_text
        return _Response()


class _FakeOpenAIClient:
    def __init__(self):
        self.responses = _FakeResponses()


class CriteriaExtractorServiceTests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._tmpdir = tempfile.TemporaryDirectory()
        cls.workbook_path = Path(cls._tmpdir.name) / "criteria.xlsx"
        _build_workbook(cls.workbook_path)

    @classmethod
    def tearDownClass(cls):
        cls._tmpdir.cleanup()
        super().tearDownClass()

    def test_extracts_criteria_from_all_subject_sheets_except_tutor(self):
        rows = extract_raw_criteria_from_workbook(str(self.workbook_path), class_code="5A")

        self.assertEqual(len(rows), 3)
        self.assertEqual(
            [r["criterion_text"] for r in rows],
            ["Критерий 1", "Критерий 2", "Критерий 3"],
        )

        for row in rows:
            self.assertEqual(row["class_code"], "5A")
            self.assertEqual(row["subject_name"], "Math")
            self.assertEqual(row["teacher_name"], "Ms. Frizzle")
            self.assertEqual(row["module_number"], 2)
            self.assertEqual(row["source_sheet_name"], "Math")
            self.assertTrue(row["source_workbook"].endswith("criteria.xlsx"))

    def test_missing_or_invalid_workbook_raises_predictable_exception(self):
        with self.assertRaises(WorkbookReadError):
            extract_raw_criteria_from_workbook("not_existing_workbook.xlsx", class_code="5A")

    def test_add_ai_normalized_criteria_keeps_source_and_adds_ai_field(self):
        fake_client = _FakeOpenAIClient()
        extracted_rows = [
            {
                "class_code": "5A",
                "subject_name": "Math",
                "teacher_name": "Ms. Frizzle",
                "module_number": 2,
                "criterion_text": "Сложная и двусмысленная формулировка",
                "source_sheet_name": "Math",
                "source_workbook": "criteria.xlsx",
            },
            {
                "class_code": "5A",
                "subject_name": "Math",
                "teacher_name": "Ms. Frizzle",
                "module_number": 2,
                "criterion_text": "Еще один критерий",
                "source_sheet_name": "Math",
                "source_workbook": "criteria.xlsx",
            },
        ]

        result = add_ai_normalized_criteria(extracted_rows, client=fake_client)

        self.assertEqual(len(result), 2)
        self.assertEqual(
            [row["criterion_text_ai"] for row in result],
            ["Нормализованный критерий", "Нормализованный критерий"],
        )
        self.assertEqual(
            [row["criterion_text"] for row in result],
            ["Сложная и двусмысленная формулировка", "Еще один критерий"],
        )
        self.assertEqual(len(fake_client.responses.calls), 2)

    def test_evaluate_criterion_text_with_ai_preserves_english_language_rule(self):
        fake_client = _FakeOpenAIClient()
        fake_client.responses.output_text = (
            '{"verdict":"valid","why":"ok","fix":"-","variants":["Estimate, add, and subtract integers"]}'
        )

        result = evaluate_criterion_text_with_ai("Estimate, add, and subtract integers", client=fake_client)

        self.assertEqual(result["variants"], ["Estimate, add, and subtract integers"])
        self.assertIn("Английский язык сам по себе не является ошибкой", _CRITERION_EVALUATION_PROMPT)
        self.assertIn("valid_criterion_examples", _CRITERION_EVALUATION_PROMPT)
        call = fake_client.responses.calls[0]
        self.assertIn("variants должны сохранять язык исходного критерия", call["input"][0]["content"])
        self.assertIn("Estimate, add, and subtract integers", call["input"][1]["content"])

    def test_evaluate_criterion_text_with_ai_sends_active_whitelist_examples(self):
        ValidCriterionTemplate.objects.create(
            name="Estimate, add, subtract, multiply and divide integers",
            keep_reason="Action-list criterion is measurable.",
        )
        ValidCriterionTemplate.objects.create(
            name="Inactive criterion",
            keep_reason="Should not be sent.",
            is_active=False,
        )
        fake_client = _FakeOpenAIClient()
        fake_client.responses.output_text = (
            '{"verdict":"valid","why":"ok","fix":"-","variants":["Estimate and add integers"]}'
        )

        result = evaluate_criterion_text_with_ai("Estimate and add integers", client=fake_client)

        self.assertEqual(result["verdict"], "valid")
        payload = json.loads(fake_client.responses.calls[0]["input"][1]["content"])
        self.assertEqual(payload["criterion"], "Estimate and add integers")
        self.assertEqual(
            payload["valid_criterion_examples"],
            [
                {
                    "criterion": "Estimate, add, subtract, multiply and divide integers",
                    "keep_reason": "Action-list criterion is measurable.",
                }
            ],
        )

    def test_evaluate_criterion_text_with_ai_retries_once_on_bad_json(self):
        fake_client = _FakeOpenAIClient()
        fake_client.responses.output_text = "not json"

        first = True

        def _create(**kwargs):
            nonlocal first
            fake_client.responses.calls.append(kwargs)

            class _Response:
                output_text = ""

            if first:
                _Response.output_text = "невалидный ответ"
                first = False
            else:
                _Response.output_text = '{"verdict":"partial","why":"reason","fix":"fix","variants":["v1","v2"]}'
            return _Response()

        fake_client.responses.create = _create

        result = evaluate_criterion_text_with_ai("Критерий", client=fake_client)

        self.assertEqual(result["verdict"], "partial")
        self.assertEqual(result["variants"], ["v1", "v2"])
        self.assertEqual(len(fake_client.responses.calls), 2)

    def test_evaluate_criterion_text_with_ai_raises_after_retry_failure(self):
        fake_client = _FakeOpenAIClient()
        fake_client.responses.output_text = "not json"

        with self.assertRaises(CriterionNormalizationError):
            evaluate_criterion_text_with_ai("Критерий", client=fake_client)
