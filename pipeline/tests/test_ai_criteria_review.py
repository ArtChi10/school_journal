import tempfile
from pathlib import Path
from unittest.mock import patch

from django.test import TestCase
from openpyxl import Workbook

from jobs.models import JobRun
from journal_links.models import ClassSheetLink
from pipeline.models import ValidCriterionTemplate
from pipeline.ai_criteria_review import (
    AICriteriaReviewFormatError,
    AI_CRITERIA_REVIEW_PROMPT,
    JOB_TYPE,
    _request_ai_class_criteria_review,
    collect_ai_reviewable_criteria_from_workbook,
    parse_ai_class_criteria_response,
    run_ai_criteria_class_review_job,
)


def _build_ai_review_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Math"
    ws.cell(row=2, column=2, value="Учитель | Teacher")
    ws.cell(row=2, column=3, value="Teacher A")
    ws.cell(row=3, column=2, value="Модуль | Module")
    ws.cell(row=3, column=3, value="2")
    ws.cell(row=5, column=2, value="Критерии оценивания | Assessment criteria")
    ws.cell(row=5, column=3, value="Solves equations")
    ws.cell(row=5, column=4, value=1)
    ws.cell(row=5, column=5, value="")
    ws.cell(row=5, column=6, value="Комментарий")
    ws.cell(row=5, column=7, value="Ignored after comment")

    ws_science = wb.create_sheet("Science")
    ws_science.cell(row=2, column=2, value="Учитель | Teacher")
    ws_science.cell(row=2, column=3, value="Teacher B")
    ws_science.cell(row=3, column=2, value="Модуль | Module")
    ws_science.cell(row=3, column=3, value="3")
    ws_science.cell(row=5, column=2, value="Критерии оценивания | Assessment criteria")
    ws_science.cell(row=5, column=3, value="Explains experiment result")

    wb.create_sheet("Тьютор | Tutor")
    wb.create_sheet("Service")
    wb.save(path)


class _FakeResponses:
    def __init__(self):
        self.calls = []

    def create(self, **kwargs):
        self.calls.append(kwargs)

        class _Response:
            output_text = '{"criteria":[{"index":1,"verdict":"ok","reason":"clear","suggested_rewrite":""}]}'

        return _Response()


class _FakeAIClient:
    def __init__(self):
        self.responses = _FakeResponses()


class AICriteriaCollectionTests(TestCase):
    def test_collects_context_and_skips_empty_numeric_tutor_service(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "criteria.xlsx"
            _build_ai_review_workbook(workbook_path)

            result = collect_ai_reviewable_criteria_from_workbook(
                str(workbook_path),
                class_code="7A",
                sheet_url="https://docs.google.com/spreadsheets/d/source/edit",
            )

        rows = result["rows"]
        self.assertEqual(len(rows), 2)
        self.assertEqual(result["summary"]["criteria_skipped_empty"], 1)
        self.assertEqual(result["summary"]["criteria_skipped_numeric"], 1)
        self.assertEqual(result["summary"]["sheets_skipped"], 2)
        self.assertEqual(rows[0]["class_code"], "7A")
        self.assertEqual(rows[0]["subject_name"], "Math")
        self.assertEqual(rows[0]["teacher_name"], "Teacher A")
        self.assertEqual(rows[0]["module_number"], 2)
        self.assertEqual(rows[0]["criterion_text"], "Solves equations")
        self.assertEqual(rows[0]["sheet_url"], "https://docs.google.com/spreadsheets/d/source/edit")


class AICriteriaPromptTests(TestCase):
    def test_prompt_allows_english_criteria_without_translation(self):
        self.assertIn("Английский язык сам по себе не является проблемой", AI_CRITERIA_REVIEW_PROMPT)
        self.assertIn("Не переводи английский критерий на русский", AI_CRITERIA_REVIEW_PROMPT)
        self.assertIn("base verb/action-list", AI_CRITERIA_REVIEW_PROMPT)

    def test_request_sends_language_rule_with_english_criterion(self):
        fake_client = _FakeAIClient()

        result = _request_ai_class_criteria_review(
            [
                {
                    "class_code": "7A",
                    "subject_name": "Math",
                    "teacher_name": "Teacher A",
                    "module_number": 2,
                    "criterion_text": "Estimate, add, and subtract integers",
                }
            ],
            ai_client=fake_client,
            model_name="test-model",
        )

        self.assertIn('"verdict":"ok"', result)
        call = fake_client.responses.calls[0]
        self.assertIn("Английский язык сам по себе не является проблемой", call["input"][0]["content"])
        self.assertIn("Estimate, add, and subtract integers", call["input"][1]["content"])


class AICriteriaResponseParsingTests(TestCase):
    def test_parses_strict_json_response(self):
        parsed = parse_ai_class_criteria_response(
            '{"criteria":[{"index":1,"verdict":"ok","reason":"good","suggested_rewrite":""}]}',
            expected_count=1,
        )

        self.assertEqual(parsed[1]["ai_verdict"], "ok")
        self.assertEqual(parsed[1]["ai_reason"], "good")

    def test_rejects_invalid_json_response(self):
        with self.assertRaises(AICriteriaReviewFormatError):
            parse_ai_class_criteria_response("not json", expected_count=1)


class AICriteriaJobTests(TestCase):
    def setUp(self):
        self.link = ClassSheetLink.objects.create(
            class_code="7A",
            subject_name="",
            teacher_name="",
            google_sheet_url="https://docs.google.com/spreadsheets/d/source/edit",
            is_active=True,
        )

    def test_job_sends_one_batch_request_per_class_and_saves_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "criteria.xlsx"
            _build_ai_review_workbook(workbook_path)

            def _fake_ai(criteria, **_kwargs):
                self.assertEqual(len(criteria), 2)
                self.assertEqual(criteria[0]["criterion_text"], "Solves equations")
                return {
                    1: {"ai_verdict": "ok", "ai_reason": "clear", "ai_suggested_rewrite": ""},
                    2: {"ai_verdict": "problem", "ai_reason": "too broad", "ai_suggested_rewrite": "Explains the experiment result using evidence."},
                }

            with (
                patch("pipeline.ai_criteria_review.fetch_workbook_for_link", return_value=workbook_path),
                patch("pipeline.ai_criteria_review.evaluate_class_criteria_with_ai", side_effect=_fake_ai) as ai_mock,
            ):
                job_run = run_ai_criteria_class_review_job(class_code="7A")

        self.assertEqual(job_run.job_type, JOB_TYPE)
        self.assertEqual(job_run.status, JobRun.Status.SUCCESS)
        self.assertEqual(ai_mock.call_count, 1)
        summary = job_run.result_json["summary"]
        self.assertEqual(summary["classes_checked"], 1)
        self.assertEqual(summary["criteria_sent_to_ai"], 2)
        self.assertEqual(summary["criteria_skipped_empty"], 1)
        self.assertEqual(summary["criteria_skipped_numeric"], 1)
        self.assertEqual(summary["criteria_ok"], 1)
        self.assertEqual(summary["criteria_problem"], 1)
        self.assertEqual(summary["ai_requests_total"], 1)
        self.assertEqual(summary["ai_requests_failed"], 0)
        self.assertEqual(job_run.result_json["report"]["status"], "not_configured")
        self.assertEqual(len(job_run.result_json["rows"]), 2)
        self.assertTrue(job_run.logs.filter(message="AI batch request started").exists())
        self.assertTrue(job_run.logs.filter(message="AI criteria review finished").exists())

    def test_active_whitelist_overrides_problem_verdict_after_ai_response(self):
        ValidCriterionTemplate.objects.create(
            name="Explains   experiment result",
            keep_reason="Это наблюдаемый и проверяемый результат.",
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "criteria.xlsx"
            _build_ai_review_workbook(workbook_path)

            def _fake_ai(criteria, **_kwargs):
                self.assertEqual(len(criteria), 2)
                return {
                    1: {"ai_verdict": "ok", "ai_reason": "clear", "ai_suggested_rewrite": ""},
                    2: {
                        "ai_verdict": "problem",
                        "ai_reason": "too broad",
                        "ai_suggested_rewrite": "Translate or rewrite.",
                    },
                }

            with (
                patch("pipeline.ai_criteria_review.fetch_workbook_for_link", return_value=workbook_path),
                patch("pipeline.ai_criteria_review.evaluate_class_criteria_with_ai", side_effect=_fake_ai),
            ):
                job_run = run_ai_criteria_class_review_job(class_code="7A")

        summary = job_run.result_json["summary"]
        self.assertEqual(summary["criteria_ok"], 2)
        self.assertEqual(summary["criteria_problem"], 0)
        self.assertEqual(summary["criteria_whitelist_overrides"], 1)

        overridden = next(row for row in job_run.result_json["rows"] if row["criterion_text"] == "Explains experiment result")
        self.assertEqual(overridden["ai_verdict"], "ok")
        self.assertTrue(overridden["ai_whitelist_override"])
        self.assertEqual(overridden["ai_original_verdict"], "problem")
        self.assertEqual(overridden["ai_original_reason"], "too broad")
        self.assertEqual(overridden["ai_original_suggested_rewrite"], "Translate or rewrite.")
        self.assertEqual(overridden["ai_suggested_rewrite"], "")
        self.assertIn("Критерий подтвержден whitelist", overridden["ai_reason"])
        self.assertIn("Это наблюдаемый и проверяемый результат", overridden["ai_reason"])

    def test_invalid_ai_json_marks_job_partial_without_losing_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "criteria.xlsx"
            _build_ai_review_workbook(workbook_path)

            with (
                patch("pipeline.ai_criteria_review.fetch_workbook_for_link", return_value=workbook_path),
                patch("pipeline.ai_criteria_review.evaluate_class_criteria_with_ai", side_effect=AICriteriaReviewFormatError("bad json")),
            ):
                job_run = run_ai_criteria_class_review_job(class_code="7A")

        self.assertEqual(job_run.status, JobRun.Status.PARTIAL)
        self.assertEqual(job_run.result_json["summary"]["ai_requests_failed"], 1)
        self.assertEqual(len(job_run.result_json["rows"]), 2)
        self.assertTrue(all(row["ai_verdict"] == "problem" for row in job_run.result_json["rows"]))
        self.assertTrue(job_run.logs.filter(message="AI batch request failed").exists())
