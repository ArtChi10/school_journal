import tempfile
from pathlib import Path
from unittest.mock import patch

from django.core.management import call_command
from django.test import TestCase
from openpyxl import Workbook

from jobs.models import JobLog, JobRun
from journal_links.models import ClassSheetLink
from pipeline.job_runner import run_build_criteria_job
from pipeline.services import CriterionNormalizationError
from pipeline.models import CriterionEntry, ValidCriterionTemplate


def _build_workbook(path: Path) -> None:
    wb = Workbook()

    ws_math = wb.active
    ws_math.title = "Math"
    ws_math["C2"] = "Ms. Frizzle"
    ws_math["C3"] = "Module 2"
    ws_math.cell(row=5, column=2, value="Критерии оценивания | \nAssessment criteria")
    ws_math.cell(row=5, column=3, value="Итоговая работа")
    ws_math.cell(row=5, column=4, value="Критерий 2")

    ws_history = wb.create_sheet("History")
    ws_history["C2"] = "Mr. History"
    ws_history["C3"] = "3"
    ws_history.cell(row=5, column=2, value="Критерии оценивания | \nAssessment criteria")
    ws_history.cell(row=5, column=3, value="Критерий H1")

    wb.save(path)


class BuildCriteriaJobTests(TestCase):
    def test_run_build_criteria_job_creates_job_logs_and_summary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "criteria.xlsx"
            _build_workbook(workbook_path)

            ClassSheetLink.objects.create(
                class_code="4A",
                subject_name="Math",
                teacher_name="Teacher A",
                google_sheet_url="https://docs.google.com/spreadsheets/d/test-id/edit#gid=0",
                is_active=True,
            )

            def _fake_ai(text, **_kwargs):
                if text == "Критерий H1":
                    raise CriterionNormalizationError("ai failed")
                return {"verdict": "valid", "why": "ok", "fix": "-", "variants": [f"AI::{text}"]}

            with (
                patch("pipeline.job_runner.run_download_descriptors_step",
                      return_value={"downloads_total":1,"downloads_success":1,"downloads_failed":0,
                                    "files":[{"link_id":1,"status":"success","path":str(workbook_path),"size_bytes":123}]}),
                patch("pipeline.job_runner.evaluate_criterion_text_with_ai", side_effect=_fake_ai),
            ):
                job = run_build_criteria_job(class_code="4A")

        self.assertEqual(job.job_type, "build_criteria_table")
        self.assertEqual(job.status, JobRun.Status.PARTIAL)
        self.assertTrue(JobLog.objects.filter(job_run=job).exists())

        summary = job.result_json["summary"]
        self.assertEqual(summary["total_sheets"], 2)
        self.assertEqual(summary["total_criteria"], 3)
        self.assertEqual(summary["ai_ok"], 2)
        self.assertEqual(summary["ai_failed"], 1)

        saved_rows = CriterionEntry.objects.filter(class_code="4A")
        self.assertEqual(saved_rows.count(), 3)
        self.assertEqual(saved_rows.exclude(criterion_text_ai="").count(), 2)
        self.assertEqual(saved_rows.filter(validation_status=CriterionEntry.ValidationStatus.VALID).count(), 2)
        self.assertEqual(saved_rows.filter(validation_status=CriterionEntry.ValidationStatus.INVALID).count(), 1)
        self.assertTrue(saved_rows.filter(criterion_text="Критерий H1", ai_verdict="failed").exists())

    def test_management_command_supports_class_code(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "criteria.xlsx"
            _build_workbook(workbook_path)

            ClassSheetLink.objects.create(
                class_code="4A",
                subject_name="Math",
                teacher_name="Teacher A",
                google_sheet_url="https://docs.google.com/spreadsheets/d/test-id/edit#gid=0",
                is_active=True,
            )

            with (
                patch("pipeline.job_runner.run_download_descriptors_step",
                      return_value={"downloads_total":1,"downloads_success":1,
                                    "downloads_failed":0,"files":[{"link_id":1,"status":"success",
                                                                   "path":str(workbook_path),"size_bytes":123}]}),
                patch("pipeline.job_runner.evaluate_criterion_text_with_ai", return_value={"verdict":"valid","why":"ok","fix":"-","variants":["AI"]}),
            ):
                call_command("build_criteria_table", "--class-code", "4A")

        self.assertTrue(JobRun.objects.filter(job_type="build_criteria_table").exists())

    def test_whitelist_criterion_is_marked_valid_without_ai(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "criteria.xlsx"
            _build_workbook(workbook_path)

            ClassSheetLink.objects.create(
                class_code="4A",
                subject_name="Math",
                teacher_name="Teacher A",
                google_sheet_url="https://docs.google.com/spreadsheets/d/test-id/edit#gid=0",
                is_active=True,
            )
            ValidCriterionTemplate.objects.create(name="Итоговая   работа")

            with patch(
                    "pipeline.job_runner.run_download_descriptors_step",
                    return_value={
                        "downloads_total": 1,
                        "downloads_success": 1,
                        "downloads_failed": 0,
                        "files": [{"link_id": 1, "status": "success", "path": str(workbook_path), "size_bytes": 123}],
                    },
            ), patch("pipeline.job_runner.evaluate_criterion_text_with_ai", return_value={"verdict":"valid","why":"ok","fix":"-","variants":["AI"]}) as ai_mock:
                job = run_build_criteria_job(class_code="4A")

        self.assertEqual(job.status, JobRun.Status.SUCCESS)
        self.assertEqual(ai_mock.call_count, 2)
        self.assertTrue(
            JobLog.objects.filter(
                job_run=job,
                message="Criterion validated by whitelist",
                context_json__reason="whitelist",
            ).exists()
        )
        self.assertTrue(CriterionEntry.objects.filter(criterion_text="Итоговая работа",
                                                      criterion_text_ai="Итоговая работа").exists())
        self.assertTrue(
            CriterionEntry.objects.filter(
                criterion_text="Итоговая работа",
                validation_status=CriterionEntry.ValidationStatus.OVERRIDE,
            ).exists()
        )

    def test_disabled_whitelist_template_sends_criterion_to_ai(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "criteria.xlsx"
            _build_workbook(workbook_path)

            ClassSheetLink.objects.create(
                class_code="4A",
                subject_name="Math",
                teacher_name="Teacher A",
                google_sheet_url="https://docs.google.com/spreadsheets/d/test-id/edit#gid=0",
                is_active=True,
            )
            ValidCriterionTemplate.objects.create(name="Итоговая работа", is_active=False)

            with patch(
                    "pipeline.job_runner.run_download_descriptors_step",
                    return_value={
                        "downloads_total": 1,
                        "downloads_success": 1,
                        "downloads_failed": 0,
                        "files": [{"link_id": 1, "status": "success", "path": str(workbook_path), "size_bytes": 123}],
                    },
            ), patch("pipeline.job_runner.evaluate_criterion_text_with_ai", return_value={"verdict":"valid","why":"ok","fix":"-","variants":["AI"]}) as ai_mock:
                run_build_criteria_job(class_code="4A")

        self.assertEqual(ai_mock.call_count, 3)

    def test_new_criterion_entry_defaults_to_pending(self):
        entry = CriterionEntry.objects.create(
            class_code="7A",
            subject_name="Physics",
            teacher_name="Ms. Photon",
            module_number=1,
            criterion_text="Критерий P1",
            source_sheet_name="Physics",
            source_workbook="criteria.xlsx",
        )
        self.assertEqual(entry.validation_status, CriterionEntry.ValidationStatus.PENDING)